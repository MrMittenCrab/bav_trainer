"""Runtime semantic component map — single source of truth for trainer coordinates."""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from .component_catalog import ComponentSpec


@dataclass
class ResolvedComponent:
    """A trainer component with workbook coordinates resolved at build time."""

    id: str
    order: int
    title: str
    short_hint: str
    semantic_key: str
    category: str
    tab: str
    cell: str
    formula: str
    expected_value: float | str | None
    tolerance: float
    depends_on: list[str] = field(default_factory=list)
    hints: list[str] = field(default_factory=list)
    related_cells: list[str] = field(default_factory=list)
    status: str = "pending"
    family_id: str = ""
    family_order: int = 0
    period_index: int | None = None
    period_end: str = ""

    @property
    def address(self) -> str:
        return f"{self.tab}!{self.cell}"

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class SemanticMap:
    """Registers and validates trainer components during workbook construction."""

    def __init__(self, expected_specs: tuple[ComponentSpec, ...] = ()) -> None:
        self._expected_specs = tuple(expected_specs)
        self._components: dict[str, ResolvedComponent] = {}
        self._by_semantic: dict[str, str] = {}
        self.rowmap: dict[str, Any] = {}

    def register(
        self,
        spec: ComponentSpec,
        tab: str,
        row: int,
        col: int,
        formula: str,
        expected_value: float | str | None,
        *,
        related_cells: list[str] | None = None,
    ) -> None:
        from openpyxl.utils import get_column_letter

        cell = f"{get_column_letter(col)}{row}"
        if spec.id in self._components:
            raise ValueError(f"Duplicate component registration: {spec.id}")
        if spec.semantic_key in self._by_semantic:
            raise ValueError(f"Duplicate semantic key: {spec.semantic_key}")

        if not formula or not str(formula).startswith("="):
            raise ValueError(
                f"Component {spec.id} ({spec.semantic_key}) requires a formula at {tab}!{cell}"
            )

        family_id = getattr(spec, "family_id", None) or spec.id
        family_order = getattr(spec, "family_order", None)
        if family_order is None:
            family_order = spec.order
        period_index = getattr(spec, "period_index", None)
        period_end = getattr(spec, "period_end", "") or ""

        self._components[spec.id] = ResolvedComponent(
            id=spec.id,
            order=spec.order,
            title=spec.title,
            short_hint=spec.short_hint,
            semantic_key=spec.semantic_key,
            category=spec.category,
            tab=tab,
            cell=cell,
            formula=str(formula),
            expected_value=expected_value,
            tolerance=spec.tolerance,
            depends_on=list(spec.depends_on),
            hints=list(spec.hints),
            related_cells=related_cells or [],
            family_id=family_id,
            family_order=int(family_order),
            period_index=period_index,
            period_end=period_end,
        )
        self._by_semantic[spec.semantic_key] = spec.id
        self.rowmap[spec.semantic_key] = {"tab": tab, "cell": cell, "row": row, "col": col}

    def get(self, component_id: str) -> ResolvedComponent:
        if component_id not in self._components:
            raise KeyError(f"Unknown component: {component_id}")
        return self._components[component_id]

    def all_ordered(self) -> list[ResolvedComponent]:
        return sorted(self._components.values(), key=lambda c: c.order)

    def validate_complete(self) -> list[str]:
        """Return blocking errors if expected concrete specs are missing or invalid."""
        errors: list[str] = []
        expected = self._expected_specs
        for spec in expected:
            if spec.id not in self._components:
                errors.append(f"Missing component: {spec.id} ({spec.semantic_key})")
                continue
            resolved = self._components[spec.id]
            if not resolved.formula.startswith("="):
                errors.append(f"{spec.id}: expected formula, got {resolved.formula!r}")
            if resolved.expected_value is None:
                errors.append(f"{spec.id}: missing expected_value (build-time computation failed)")
            for dep in spec.depends_on:
                if dep not in self._components:
                    errors.append(f"{spec.id}: dependency {dep} not registered")
        for comp in self.all_ordered():
            for dep in comp.depends_on:
                dep_comp = self._components.get(dep)
                if dep_comp and dep_comp.order >= comp.order:
                    errors.append(
                        f"{comp.id}: dependency {dep} must come before order {comp.order}"
                    )
        return errors

    def save_json(self, path: Path) -> None:
        payload = {
            "version": 1,
            "components": [c.to_dict() for c in self.all_ordered()],
            "rowmap": self.rowmap,
        }
        path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")

    @classmethod
    def _component_from_raw(cls, raw: dict[str, Any]) -> ResolvedComponent:
        cid = raw["id"]
        order = int(raw["order"])
        return ResolvedComponent(
            id=cid,
            order=order,
            title=raw["title"],
            short_hint=raw["short_hint"],
            semantic_key=raw["semantic_key"],
            category=raw["category"],
            tab=raw["tab"],
            cell=raw["cell"],
            formula=raw["formula"],
            expected_value=raw.get("expected_value"),
            tolerance=float(raw.get("tolerance", 0.01)),
            depends_on=list(raw.get("depends_on") or []),
            hints=list(raw.get("hints") or []),
            related_cells=list(raw.get("related_cells") or []),
            status=raw.get("status", "pending"),
            family_id=raw.get("family_id") or cid,
            family_order=int(raw.get("family_order") if raw.get("family_order") is not None else order),
            period_index=raw.get("period_index"),
            period_end=raw.get("period_end") or "",
        )

    @classmethod
    def load_json(cls, path: Path) -> SemanticMap:
        data = json.loads(path.read_text(encoding="utf-8"))
        smap = cls()
        smap.rowmap = data.get("rowmap", {})
        for raw in data.get("components", []):
            comp = cls._component_from_raw(raw)
            smap._components[comp.id] = comp
            smap._by_semantic[comp.semantic_key] = comp.id
        return smap

    @classmethod
    def from_workbook(cls, wb_path: Path) -> SemanticMap:
        """Load from embedded _ComponentMap sheet or sidecar JSON."""
        sidecar = wb_path.with_suffix(".component_map.json")
        if sidecar.exists():
            return cls.load_json(sidecar)
        from openpyxl import load_workbook

        wb = load_workbook(wb_path, data_only=False)
        if "_ComponentMap" not in wb.sheetnames:
            wb.close()
            raise FileNotFoundError(f"No component map in {wb_path}")
        ws = wb["_ComponentMap"]
        smap = cls()
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
            if not row.get("id"):
                continue
            hints_raw = row.get("hints", "") or ""
            hints = hints_raw.split("|") if hints_raw else []
            deps_raw = row.get("depends_on", "") or ""
            deps = deps_raw.split(",") if deps_raw else []
            related_raw = row.get("related_cells", "") or ""
            related = related_raw.split(",") if related_raw else []
            period_index = row.get("period_index")
            if period_index is not None and period_index != "":
                period_index = int(period_index)
            else:
                period_index = None
            raw = {
                "id": row["id"],
                "order": int(row["order"]),
                "title": row["title"],
                "short_hint": row["short_hint"],
                "semantic_key": row["semantic_key"],
                "category": row["category"],
                "tab": row["tab"],
                "cell": row["cell"],
                "formula": row["formula"],
                "expected_value": row.get("expected_value"),
                "tolerance": float(row.get("tolerance", 0.01)),
                "depends_on": deps,
                "hints": hints,
                "related_cells": related,
                "status": row.get("status", "pending"),
                "family_id": row.get("family_id") or row["id"],
                "family_order": row.get("family_order"),
                "period_index": period_index,
                "period_end": row.get("period_end") or "",
            }
            comp = cls._component_from_raw(raw)
            smap._components[comp.id] = comp
            smap._by_semantic[comp.semantic_key] = comp.id
        wb.close()
        return smap
