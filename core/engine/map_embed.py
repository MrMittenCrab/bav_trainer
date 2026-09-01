"""Embed semantic component map into a workbook."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from .semantic_map import SemanticMap

COMPONENT_MAP_SHEET = "_ComponentMap"


def embed_component_map_sheet(wb: Workbook, smap: SemanticMap) -> None:
    if COMPONENT_MAP_SHEET in wb.sheetnames:
        del wb[COMPONENT_MAP_SHEET]
    ws = wb.create_sheet(COMPONENT_MAP_SHEET)
    ws.sheet_state = "hidden"
    headers = [
        "id", "order", "title", "short_hint", "semantic_key", "category",
        "tab", "cell", "formula", "expected_value", "tolerance",
        "depends_on", "hints", "related_cells", "status",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h).font = Font(bold=True)
    for i, comp in enumerate(smap.all_ordered(), start=2):
        ws.cell(row=i, column=1, value=comp.id)
        ws.cell(row=i, column=2, value=comp.order)
        ws.cell(row=i, column=3, value=comp.title)
        ws.cell(row=i, column=4, value=comp.short_hint)
        ws.cell(row=i, column=5, value=comp.semantic_key)
        ws.cell(row=i, column=6, value=comp.category)
        ws.cell(row=i, column=7, value=comp.tab)
        ws.cell(row=i, column=8, value=comp.cell)
        ws.cell(row=i, column=9, value=comp.formula)
        ws.cell(row=i, column=10, value=comp.expected_value)
        ws.cell(row=i, column=11, value=comp.tolerance)
        ws.cell(row=i, column=12, value=",".join(comp.depends_on))
        ws.cell(row=i, column=13, value="|".join(comp.hints))
        ws.cell(row=i, column=14, value=",".join(comp.related_cells))
        ws.cell(row=i, column=15, value=comp.status)
