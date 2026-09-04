"""Build the complete reference BAV workbook and semantic component map."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..data.interface import LineItem, StandardizedFinancials
from ..data.line_identity import line_identity
from ..model.classification import BALANCE_SHEET_CATEGORIES
from ..model.financial_math import compute_anchor
from ..model.line_resolver import resolve_line, workbook_row_for
from ..model.ri_engine import run_scenario, weighted_ivps
from .component_catalog import DEFERRED_COMPONENT_SPECS, expand_historical_specs
from .map_embed import embed_component_map_sheet
from .semantic_map import SemanticMap

NUM_FMT = "#,##0;(#,##0)"
PCT_FMT = "0.0%"
SOURCE_START_ROW = 7
BLUE = Font(color="0000FF")
BOLD = Font(bold=True)
ORANGE = PatternFill("solid", start_color="FCE5CD")
YELLOW = PatternFill("solid", start_color="FFF2CC")
GREEN = PatternFill("solid", start_color="D9EAD3")

DEFERRED_TAB_NAMES = ("Model_Bear", "Model_Base", "Model_Bull", "Scenario_Summary")
DEFERRED_PLACEHOLDER = "Deferred from historical-only v1"


class ReferenceModelBuilder:
    """Construct reference workbook and populate SemanticMap at build time."""

    def __init__(
        self,
        financials: StandardizedFinancials,
        assumptions: dict[str, Any] | None = None,
        *,
        include_deferred_forecast: bool = False,
    ):
        self.fin = financials
        self.periods = financials.fiscal_years() or financials.period_dates()
        self.include_deferred_forecast = include_deferred_forecast
        self.assumptions = dict(assumptions or {})
        self.assumptions.setdefault("classificationOverrides", {})
        self.rowmap: dict[str, Any] = {}
        self.historical_specs = expand_historical_specs(self.periods)
        self.semantic_map = SemanticMap(expected_specs=self.historical_specs)
        self._historical_spec_index = {
            (s.family_id, s.period_index): s for s in self.historical_specs
        }
        self._deferred_spec_index = {c.id: c for c in DEFERRED_COMPONENT_SPECS}
        overrides = self.assumptions.get("classificationOverrides") or {}
        self.anchor = compute_anchor(
            financials,
            self.periods,
            classification_overrides=overrides,
        )
        self._n = len(self.periods)
        self._last_fy_col = 2 + self._n - 1
        self._first_fc_col = 2 + self._n
        self._scenario_results: dict[str, Any] = {}
        self._base_result = None

        if self.include_deferred_forecast:
            # Internal/legacy path only — never enabled by normal v1 build.
            if "scenarios" not in self.assumptions or "marketData" not in self.assumptions:
                defaults = self._default_assumptions()
                for key, value in defaults.items():
                    self.assumptions.setdefault(key, value)
            shares = self.assumptions["marketData"]["dilutedShares"]
            self._scenario_results = {
                name: run_scenario(
                    self.assumptions["scenarios"][name],
                    self.anchor,
                    shares,
                )
                for name in ("Bear", "Base", "Bull")
            }
            self._base_result = self._scenario_results["Base"]

    def _default_assumptions(self) -> dict[str, Any]:
        anchor_rev = 1000.0
        if self.fin.income_statement and self.periods:
            rev_item = resolve_line(
                self.fin.income_statement, "revenue", required=True
            ).item
            assert rev_item is not None
            anchor_rev = rev_item.values.get(self.periods[-1]) or 1000.0
        growth = [0.10] * 10
        margin = [0.15] * 10
        nowc = [0.05] * 10
        nola = [0.50] * 10
        return {
            "schemaVersion": 2,
            "ticker": self.fin.ticker,
            "company": self.fin.company_name,
            "marketData": {
                "price": 50.0,
                "priceDate": date.today().isoformat(),
                "dilutedShares": 1000.0,
                "riskFreeRate": 0.04,
                "equityRiskPremium": 0.05,
            },
            "scenarios": {
                name: {
                    "probability": prob,
                    "beta": beta,
                    "costOfEquity": 0.04 + beta * 0.05,
                    "taxRate": 0.165,
                    "terminalGrowth": 0.03,
                    "growthVector": growth,
                    "marginVector": margin,
                    "nowcRatioVector": nowc,
                    "nolaRatioVector": nola,
                }
                for name, prob, beta in (
                    ("Bear", 0.25, 1.3),
                    ("Base", 0.50, 1.15),
                    ("Bull", 0.25, 1.05),
                )
            },
            "classificationOverrides": {},
            "meta": {"anchorRevenue": anchor_rev, "currency": self.fin.units},
        }

    def build(self, output_path: Path) -> SemanticMap:
        wb = Workbook()
        self._build_source_tabs(wb)
        self._build_condensed(wb)
        self._build_dupont(wb)
        if self.include_deferred_forecast:
            for scenario in ("Bear", "Base", "Bull"):
                self._build_model_tab(wb, scenario)
            self._build_scenario_summary(wb)
        else:
            for name in DEFERRED_TAB_NAMES:
                ws = wb.create_sheet(name)
                ws["A1"] = DEFERRED_PLACEHOLDER
                ws.sheet_state = "hidden"

        errors = self.semantic_map.validate_complete()
        if errors:
            raise ValueError("Component map validation failed:\n" + "\n".join(errors))

        embed_component_map_sheet(wb, self.semantic_map)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        sidecar = output_path.with_suffix(".component_map.json")
        self.semantic_map.save_json(sidecar)
        assumptions_path = output_path.with_suffix(".assumptions.json")
        assumptions_path.write_text(json.dumps(self.assumptions, indent=2) + "\n", encoding="utf-8")
        rowmap_path = output_path.parent / "rowmap.json"
        rowmap_path.write_text(
            json.dumps({**self.rowmap, **self.semantic_map.rowmap}, indent=2) + "\n",
            encoding="utf-8",
        )
        return self.semantic_map

    def _col(self, idx: int) -> str:
        return get_column_letter(idx)

    def _register_historical(
        self,
        family_id: str,
        period_index: int,
        tab: str,
        row: int,
        col: int,
        formula: str,
        expected: float | str,
        related: list[str] | None = None,
    ) -> None:
        spec = self._historical_spec_index[(family_id, period_index)]
        self.semantic_map.register(
            spec, tab, row, col, formula, expected, related_cells=related
        )

    def _register_deferred(
        self,
        spec_id: str,
        tab: str,
        row: int,
        col: int,
        formula: str,
        expected: float | str,
        related: list[str] | None = None,
    ) -> None:
        if spec_id not in self._deferred_spec_index:
            raise KeyError(f"Unknown deferred component spec: {spec_id}")
        spec = self._deferred_spec_index[spec_id]
        self.semantic_map.register(
            spec, tab, row, col, formula, expected, related_cells=related
        )

    def _resolved_source_row(self, items: list[LineItem], concept: str, *, required: bool = False) -> int | None:
        """Workbook row for a canonical concept — same LineItem as compute_anchor()."""
        return workbook_row_for(
            resolve_line(items, concept, required=required),
            start_row=SOURCE_START_ROW,
        )

    def _header_block(self, ws, statement: str) -> None:
        ws["A1"] = f"Company: {self.fin.company_name} ({self.fin.ticker})"
        ws["A2"] = f"Statement: {statement}"
        ws["A3"] = f"Units: {self.fin.units}"
        ws["A4"] = f"Source: manual ingestion ({self.fin.jurisdiction})"
        ws["A6"] = "Line Item"
        ws["A6"].font = BOLD
        for j, pd in enumerate(self.periods):
            c = ws.cell(row=6, column=2 + j, value=pd)
            c.number_format = "mmm dd, yyyy"
            c.font = BOLD
        ws.column_dimensions["A"].width = 48
        for j in range(self._n):
            ws.column_dimensions[self._col(2 + j)].width = 16

    def _fill_statement(self, ws, items: list[LineItem], start_row: int = 7) -> int:
        r = start_row
        sheet = ws.title
        for item in items:
            ws.cell(row=r, column=1, value=item.label)
            for j, pd in enumerate(self.periods):
                val = item.values.get(pd)
                c = ws.cell(row=r, column=2 + j, value=val)
                c.number_format = NUM_FMT
            self.rowmap[f"{sheet}!{line_identity(item).key()}"] = r
            r += 1
        return r

    def _build_source_tabs(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Income Statement"
        self._header_block(ws, "Income Statement")
        self._fill_statement(ws, self.fin.income_statement)
        ws = wb.create_sheet("Balance Sheet")
        self._header_block(ws, "Balance Sheet")
        self._fill_statement(ws, self.fin.balance_sheet)
        ws = wb.create_sheet("Cash Flow Statement")
        self._header_block(ws, "Cash Flow Statement")
        self._fill_statement(ws, self.fin.cash_flow)

    def _build_condensed(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Condensed Financials")
        ws["A1"] = f"{self.fin.company_name} ({self.fin.ticker}) — Condensed Financials"
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 32

        r = 4
        ws.cell(row=r, column=1, value="BALANCE SHEET CLASSIFICATION").font = BOLD
        r += 1
        ws.cell(row=r, column=1, value="Line Item").font = BOLD
        ws.cell(row=r, column=2, value="Classification").font = BOLD
        for j, pd in enumerate(self.periods):
            c = ws.cell(row=r, column=3 + j, value=pd)
            c.number_format = "mmm dd, yyyy"
            c.font = BOLD
            ws.column_dimensions[self._col(3 + j)].width = 14
        r += 1

        notes_col = 3 + self._n
        ws.cell(row=r - 1, column=notes_col, value="Notes").font = BOLD
        ws.column_dimensions[self._col(notes_col)].width = 42

        class_start = r
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(BALANCE_SHEET_CATEGORIES)}"',
            allow_blank=False,
        )
        ws.add_data_validation(dv)
        # Shared decisions drive defaults; SUMIF stays on-sheet for live reclassification.
        reform = self.anchor.reformulation
        for idx in reform.detail_indices:
            item = self.fin.balance_sheet[idx]
            decision = reform.decisions[idx]
            ws.cell(row=r, column=1, value=item.label)
            cat_cell = ws.cell(row=r, column=2, value=decision.category)
            dv.add(cat_cell)
            for j, pd in enumerate(self.periods):
                c = ws.cell(row=r, column=3 + j, value=item.values.get(pd))
                c.number_format = NUM_FMT
            note_parts: list[str] = []
            if decision.ambiguous:
                note_parts.append(f"⚠ Review: {decision.reason or 'judgment required'}")
            if decision.overridden:
                note_parts.append(f"Override → {decision.category}")
            if note_parts:
                ws.cell(row=r, column=notes_col, value="; ".join(note_parts))
            r += 1
        class_end = r - 1
        self.rowmap["condensed_class_start"] = class_start
        self.rowmap["condensed_class_end"] = class_end
        self.rowmap["condensed_class_value_col0"] = 3

        r += 1
        ws.cell(row=r, column=1, value="CONDENSED INCOME STATEMENT").font = BOLD
        for j, pd in enumerate(self.periods):
            c = ws.cell(row=r, column=2 + j, value=pd)
            c.number_format = "mmm dd, yyyy"
            c.font = BOLD
        r += 1

        ni_src = self._resolved_source_row(self.fin.income_statement, "net_income", required=True)
        rev_src = self._resolved_source_row(self.fin.income_statement, "revenue", required=True)
        pretax_src = self._resolved_source_row(self.fin.income_statement, "pretax_income")
        tax_src = self._resolved_source_row(self.fin.income_statement, "tax_expense")
        int_exp_src = self._resolved_source_row(self.fin.income_statement, "interest_expense")
        int_inc_src = self._resolved_source_row(self.fin.income_statement, "interest_income")
        equity_src = self._resolved_source_row(self.fin.balance_sheet, "total_equity")
        row_nums: dict[str, int] = {}
        hist = self.anchor.historical
        for label, src_row, bold, family_id, expected_series in [
            ("Revenue", rev_src, False, "revenue_link", hist.revenue),
            ("Net Income", ni_src, False, "net_income_link", hist.net_income),
            ("Pretax Income", pretax_src, False, None, None),
            ("Tax Expense", tax_src, False, None, None),
            ("Interest Expense", int_exp_src, False, None, None),
            ("Interest Income", int_inc_src, False, None, None),
        ]:
            if not src_row:
                continue
            ws.cell(row=r, column=1, value=label).font = Font(bold=bold)
            for j in range(self._n):
                col = self._col(2 + j)
                formula = f"='Income Statement'!{col}{src_row}"
                ws.cell(row=r, column=2 + j, value=formula)
                if family_id is not None and expected_series is not None:
                    self._register_historical(
                        family_id,
                        j,
                        "Condensed Financials",
                        r,
                        2 + j,
                        formula,
                        expected_series[j],
                    )
            row_nums[label] = r
            if label == "Net Income":
                self.rowmap["condensed_ni_row"] = r
            if label == "Revenue":
                self.rowmap["condensed_revenue_row"] = r
            r += 1

        etr_row = r
        ws.cell(row=r, column=1, value="Effective Tax Rate")
        for j in range(self._n):
            col = self._col(2 + j)
            if "Tax Expense" in row_nums and "Pretax Income" in row_nums:
                pretax_r = row_nums["Pretax Income"]
                tax_r = row_nums["Tax Expense"]
                formula = f"=IF({col}{pretax_r}=0,0,-{col}{tax_r}/{col}{pretax_r})"
            else:
                formula = "=0"
            ws.cell(row=r, column=2 + j, value=formula).number_format = PCT_FMT
            self._register_historical(
                "effective_tax_rate_fy",
                j,
                "Condensed Financials",
                r,
                2 + j,
                formula,
                hist.effective_tax_rate[j],
            )
        row_nums["Effective Tax Rate"] = etr_row
        r += 1

        # Net interest: missing optional interest lines treated as zero (matches Python).
        net_int_row = r
        ws.cell(row=r, column=1, value="Net Interest")
        has_ie = "Interest Expense" in row_nums
        has_ii = "Interest Income" in row_nums
        for j in range(self._n):
            col = self._col(2 + j)
            if has_ie and has_ii:
                f: str = (
                    f"=-({col}{row_nums['Interest Expense']}"
                    f"+{col}{row_nums['Interest Income']})"
                )
            elif has_ie:
                f = f"=-{col}{row_nums['Interest Expense']}"
            elif has_ii:
                f = f"=-{col}{row_nums['Interest Income']}"
            else:
                f = "=0"
            ws.cell(row=r, column=2 + j, value=f).number_format = NUM_FMT
            self._register_historical(
                "net_interest_fy",
                j,
                "Condensed Financials",
                r,
                2 + j,
                f,
                hist.net_interest[j],
            )
        row_nums["Net Interest"] = net_int_row
        r += 1

        niat_row = r
        ws.cell(row=r, column=1, value="Net Interest After Tax")
        for j in range(self._n):
            col = self._col(2 + j)
            formula = f"={col}{net_int_row}*(1-{col}{etr_row})"
            ws.cell(row=r, column=2 + j, value=formula).number_format = NUM_FMT
            self._register_historical(
                "net_interest_after_tax_fy",
                j,
                "Condensed Financials",
                r,
                2 + j,
                formula,
                hist.net_interest_after_tax[j],
            )
        row_nums["Net Interest After Tax"] = niat_row
        self.rowmap["condensed_niat_row"] = niat_row
        r += 1

        nopat_row = r
        ws.cell(row=r, column=1, value="NOPAT").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            formula = f"={col}{row_nums['Net Income']}+{col}{niat_row}"
            c = ws.cell(row=r, column=2 + j, value=formula)
            c.fill = GREEN
            c.number_format = NUM_FMT
            self._register_historical(
                "nopat_fy",
                j,
                "Condensed Financials",
                r,
                2 + j,
                formula,
                hist.nopat[j],
            )
        r += 1
        self.rowmap["condensed_nopat_row"] = nopat_row

        r += 1
        ws.cell(row=r, column=1, value="CONDENSED BALANCE SHEET").font = BOLD
        r += 1

        def _class_sumif(category: str, value_col: str) -> str:
            return (
                f'SUMIF($B${class_start}:$B${class_end},"{category}",'
                f"{value_col}${class_start}:{value_col}${class_end})"
            )

        def _fill_sumif_row(label: str, category: str, *, bold: bool = False) -> int:
            nonlocal r
            row = r
            ws.cell(row=r, column=1, value=label).font = Font(bold=bold)
            for j in range(self._n):
                vcol = self._col(3 + j)
                formula = f"={_class_sumif(category, vcol)}"
                c = ws.cell(row=r, column=2 + j, value=formula)
                c.number_format = NUM_FMT
            r += 1
            return row

        cat_totals = self.anchor.reformulation.category_totals
        reform = self.anchor.reformulation

        def _register_all_periods(
            family_id: str, row: int, expected_series: list[float]
        ) -> None:
            for j in range(self._n):
                self._register_historical(
                    family_id,
                    j,
                    "Condensed Financials",
                    row,
                    2 + j,
                    str(ws.cell(row=row, column=2 + j).value),
                    expected_series[j],
                )

        owca_row = _fill_sumif_row(
            "Operating Working Capital Assets", "Operating Working Capital Asset"
        )
        _register_all_periods(
            "owca_agg", owca_row, cat_totals["Operating Working Capital Asset"]
        )
        owcl_row = _fill_sumif_row(
            "Operating Working Capital Liabilities",
            "Operating Working Capital Liability",
        )
        _register_all_periods(
            "owcl_agg", owcl_row, cat_totals["Operating Working Capital Liability"]
        )

        nowc_row = r
        ws.cell(row=r, column=1, value="NOWC").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            c = ws.cell(row=r, column=2 + j, value=f"={col}{owca_row}-{col}{owcl_row}")
            c.number_format = NUM_FMT
        r += 1
        _register_all_periods("nowc_agg", nowc_row, list(reform.nowc))

        olta_row = _fill_sumif_row(
            "Operating Long-Term Assets", "Operating Long-Term Asset"
        )
        _register_all_periods(
            "olta_agg", olta_row, cat_totals["Operating Long-Term Asset"]
        )
        oltl_row = _fill_sumif_row(
            "Operating Long-Term Liabilities", "Operating Long-Term Liability"
        )
        _register_all_periods(
            "oltl_agg", oltl_row, cat_totals["Operating Long-Term Liability"]
        )

        nola_row = r
        ws.cell(row=r, column=1, value="NOLA").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            c = ws.cell(row=r, column=2 + j, value=f"={col}{olta_row}-{col}{oltl_row}")
            c.number_format = NUM_FMT
        r += 1
        _register_all_periods("nola_agg", nola_row, list(reform.nola))

        noa_row = r
        ws.cell(row=r, column=1, value="NOA").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            c = ws.cell(row=r, column=2 + j, value=f"={col}{nowc_row}+{col}{nola_row}")
            c.number_format = NUM_FMT
        r += 1
        _register_all_periods("noa_agg", noa_row, list(reform.noa))

        fa_row = _fill_sumif_row("Financial Assets", "Financial Asset")
        _register_all_periods("financial_assets_agg", fa_row, cat_totals["Financial Asset"])
        fl_row = _fill_sumif_row("Financial Liabilities", "Financial Liability")
        _register_all_periods(
            "financial_liabilities_agg", fl_row, cat_totals["Financial Liability"]
        )

        nd_row = r
        ws.cell(row=r, column=1, value="Net Debt").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            c = ws.cell(row=r, column=2 + j, value=f"={col}{fl_row}-{col}{fa_row}")
            c.number_format = NUM_FMT
        r += 1
        _register_all_periods("net_debt", nd_row, list(reform.net_debt))

        equity_row = r
        ws.cell(row=r, column=1, value="Equity (NOA - Net Debt)").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            c = ws.cell(row=r, column=2 + j, value=f"={col}{noa_row}-{col}{nd_row}")
            c.number_format = NUM_FMT
        r += 1
        _register_all_periods(
            "equity_reformulated_fy", equity_row, list(reform.implied_equity)
        )

        reported_eq_row = r
        ws.cell(row=r, column=1, value="Reported Equity").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            if equity_src:
                f: str | None = f"='Balance Sheet'!{col}{equity_src}"
            else:
                f = None
            c = ws.cell(row=r, column=2 + j, value=f)
            c.number_format = NUM_FMT
        r += 1

        total_cap_row = r
        ws.cell(row=r, column=1, value="Total Capital").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            c = ws.cell(row=r, column=2 + j, value=f"={col}{nd_row}+{col}{equity_row}")
            c.number_format = NUM_FMT
        r += 1

        check_row = r
        ws.cell(row=r, column=1, value="CHECK").font = BOLD
        for j in range(self._n):
            col = self._col(2 + j)
            if equity_src:
                f = (
                    f'=IF(ABS({col}{equity_row}-{col}{reported_eq_row})<1,"OK","CHECK")'
                )
            else:
                f = '="UNVERIFIED"'
            ws.cell(row=r, column=2 + j, value=f)
        r += 1

        self.rowmap["condensed_nowc_row"] = nowc_row
        self.rowmap["condensed_nola_row"] = nola_row
        self.rowmap["condensed_noa_row"] = noa_row
        self.rowmap["condensed_nd_row"] = nd_row
        self.rowmap["condensed_equity_row"] = equity_row
        self.rowmap["condensed_reported_equity_row"] = reported_eq_row
        self.rowmap["condensed_total_capital_row"] = total_cap_row
        self.rowmap["condensed_check_row"] = check_row
        self.rowmap["condensed_nola_via_noa"] = True

    def _build_dupont(self, wb: Workbook) -> None:
        ws = wb.create_sheet("ALT DuPont")
        ws["A1"] = f"{self.fin.company_name} — DuPont Decomposition"
        ws["A2"] = "ROE = RNOA + FLEV × Spread"
        ws.column_dimensions["A"].width = 42
        r_hdr = 4
        ws.cell(row=r_hdr, column=1, value="Metric").font = BOLD
        for j, pd in enumerate(self.periods):
            c = ws.cell(row=r_hdr, column=2 + j, value=pd)
            c.number_format = "mmm dd, yyyy"
            c.font = BOLD
            ws.column_dimensions[self._col(2 + j)].width = 14

        nopat_r = self.rowmap["condensed_nopat_row"]
        niat_r = self.rowmap["condensed_niat_row"]
        noa_r = self.rowmap["condensed_noa_row"]
        nd_r = self.rowmap["condensed_nd_row"]
        eq_r = self.rowmap["condensed_equity_row"]
        ni_r = self.rowmap["condensed_ni_row"]
        rev_r = self.rowmap["condensed_revenue_row"]

        metrics = [
            "Sales Growth",
            "NOPAT Margin",
            "RNOA",
            "After-tax CoD",
            "Spread",
            "FLEV",
            "ROE (decomposed)",
            "Actual ROE",
        ]
        metric_rows: dict[str, int] = {}
        r = r_hdr + 1
        for metric in metrics:
            ws.cell(row=r, column=1, value=metric)
            metric_rows[metric] = r
            r += 1

        sales_row = metric_rows["Sales Growth"]
        margin_row = metric_rows["NOPAT Margin"]
        rnoa_row = metric_rows["RNOA"]
        cod_row = metric_rows["After-tax CoD"]
        spread_row = metric_rows["Spread"]
        flev_row = metric_rows["FLEV"]
        roe_row = metric_rows["ROE (decomposed)"]
        actual_row = metric_rows["Actual ROE"]

        dup = self.anchor.dupont
        na = "N/A"

        for j in range(self._n):
            out_col_idx = 2 + j
            out_col = self._col(out_col_idx)
            src_col = self._col(2 + j)
            src_prev = self._col(2 + j - 1) if j > 0 else None

            # NOPAT Margin — all periods
            margin_f = (
                f"=IF('Condensed Financials'!{src_col}{rev_r}=0,0,"
                f"'Condensed Financials'!{src_col}{nopat_r}/"
                f"'Condensed Financials'!{src_col}{rev_r})"
            )
            cell = ws.cell(row=margin_row, column=out_col_idx, value=margin_f)
            cell.number_format = PCT_FMT
            self._register_historical(
                "nopat_margin",
                j,
                "ALT DuPont",
                margin_row,
                out_col_idx,
                margin_f,
                dup["NOPAT Margin"][j],
            )

            if j == 0:
                for row in (
                    sales_row,
                    rnoa_row,
                    cod_row,
                    spread_row,
                    flev_row,
                    roe_row,
                    actual_row,
                ):
                    ws.cell(row=row, column=out_col_idx, value=na)
                continue

            assert src_prev is not None
            sales_f = (
                f"=IF('Condensed Financials'!{src_prev}{rev_r}=0,0,"
                f"'Condensed Financials'!{src_col}{rev_r}/"
                f"'Condensed Financials'!{src_prev}{rev_r}-1)"
            )
            ws.cell(row=sales_row, column=out_col_idx, value=sales_f).number_format = PCT_FMT
            self._register_historical(
                "sales_growth",
                j,
                "ALT DuPont",
                sales_row,
                out_col_idx,
                sales_f,
                dup["Sales Growth"][j],
            )

            rnoa_f = (
                f"=IF((('Condensed Financials'!{src_col}{noa_r}+"
                f"'Condensed Financials'!{src_prev}{noa_r})/2)=0,0,"
                f"'Condensed Financials'!{src_col}{nopat_r}/"
                f"(('Condensed Financials'!{src_col}{noa_r}+"
                f"'Condensed Financials'!{src_prev}{noa_r})/2))"
            )
            ws.cell(row=rnoa_row, column=out_col_idx, value=rnoa_f).number_format = PCT_FMT
            self._register_historical(
                "rnoa",
                j,
                "ALT DuPont",
                rnoa_row,
                out_col_idx,
                rnoa_f,
                dup["RNOA"][j],
            )

            cod_f = (
                f"=IF((('Condensed Financials'!{src_col}{nd_r}+"
                f"'Condensed Financials'!{src_prev}{nd_r})/2)=0,0,"
                f"'Condensed Financials'!{src_col}{niat_r}/"
                f"(('Condensed Financials'!{src_col}{nd_r}+"
                f"'Condensed Financials'!{src_prev}{nd_r})/2))"
            )
            ws.cell(row=cod_row, column=out_col_idx, value=cod_f).number_format = PCT_FMT
            self._register_historical(
                "after_tax_cod",
                j,
                "ALT DuPont",
                cod_row,
                out_col_idx,
                cod_f,
                dup["After-tax CoD"][j],
            )

            spread_f = f"={out_col}{rnoa_row}-{out_col}{cod_row}"
            ws.cell(row=spread_row, column=out_col_idx, value=spread_f).number_format = PCT_FMT
            self._register_historical(
                "spread",
                j,
                "ALT DuPont",
                spread_row,
                out_col_idx,
                spread_f,
                dup["Spread"][j],
            )

            flev_f = (
                f"=IF((('Condensed Financials'!{src_col}{eq_r}+"
                f"'Condensed Financials'!{src_prev}{eq_r})/2)=0,0,"
                f"(('Condensed Financials'!{src_col}{nd_r}+"
                f"'Condensed Financials'!{src_prev}{nd_r})/2)/"
                f"(('Condensed Financials'!{src_col}{eq_r}+"
                f"'Condensed Financials'!{src_prev}{eq_r})/2))"
            )
            ws.cell(row=flev_row, column=out_col_idx, value=flev_f)
            self._register_historical(
                "flev",
                j,
                "ALT DuPont",
                flev_row,
                out_col_idx,
                flev_f,
                dup["FLEV"][j],
            )

            roe_f = f"={out_col}{rnoa_row}+{out_col}{flev_row}*({out_col}{spread_row})"
            ws.cell(row=roe_row, column=out_col_idx, value=roe_f).number_format = PCT_FMT
            self._register_historical(
                "roe_decomp",
                j,
                "ALT DuPont",
                roe_row,
                out_col_idx,
                roe_f,
                dup["ROE (decomposed)"][j],
            )

            actual_f = (
                f"=IF((('Condensed Financials'!{src_col}{eq_r}+"
                f"'Condensed Financials'!{src_prev}{eq_r})/2)=0,0,"
                f"'Condensed Financials'!{src_col}{ni_r}/"
                f"(('Condensed Financials'!{src_col}{eq_r}+"
                f"'Condensed Financials'!{src_prev}{eq_r})/2))"
            )
            ws.cell(row=actual_row, column=out_col_idx, value=actual_f).number_format = PCT_FMT
            self._register_historical(
                "actual_roe",
                j,
                "ALT DuPont",
                actual_row,
                out_col_idx,
                actual_f,
                dup["Actual ROE"][j],
            )

    def _build_model_tab(self, wb: Workbook, scenario: str) -> None:
        ws = wb.create_sheet(f"Model_{scenario}")
        sc = self.assumptions["scenarios"][scenario]
        result = self._scenario_results[scenario]
        ws["A1"] = f"{self.fin.company_name} — {scenario} Scenario"
        ws.freeze_panes = "B1"
        ws.column_dimensions["A"].width = 42

        ws["A5"] = "Cost of Equity (Ke)"
        ws["B5"] = sc["costOfEquity"]
        ws["B5"].font = BLUE
        ws["B5"].number_format = PCT_FMT
        ws["A6"] = "Tax rate"
        ws["B6"] = sc.get("taxRate", 0.165)
        ws["B6"].number_format = PCT_FMT
        # Historical average is already after-tax — use directly; do not tax again.
        ws["A7"] = "After-tax CoD"
        ws["B7"] = self.anchor.hist_avg_after_tax_cod
        ws["B7"].number_format = PCT_FMT
        ws["A8"] = "Financial leverage"
        ws["B8"] = self.anchor.leverage
        ws["B8"].number_format = PCT_FMT
        ws["A9"] = "Terminal growth (g)"
        ws["B9"] = sc["terminalGrowth"]
        ws["B9"].font = BLUE
        ws["B9"].number_format = PCT_FMT

        fc = self._first_fc_col
        anchor_col = self._col(self._last_fy_col)
        rev_row = self._resolved_source_row(
            self.fin.income_statement, "revenue", required=True
        )
        nowc_hist = self.rowmap["condensed_nowc_row"]
        noa_hist = self.rowmap["condensed_noa_row"]
        nd_hist = self.rowmap["condensed_nd_row"]
        # NOLA_hist = NOA - NOWC for Y1 bridge
        nola_hist_formula = (
            f"='Condensed Financials'!{anchor_col}{noa_hist}"
            f"-'Condensed Financials'!{anchor_col}{nowc_hist}"
        )

        sales_row, margin_row = 21, 22
        nowc_row, nola_row, nd_row, eq_row = 23, 24, 25, 26
        nopat_row, ni_row, ae_row = 27, 28, 29
        disc_row, pv_ae_row = 30, 31
        sum_pv_ae_row, tv_row, tv_pv_row = 35, 36, 37
        iv_row, shares_row, ivps_row = 38, 39, 40

        ws.cell(row=20, column=1, value="FORECAST BLOCK").font = BOLD
        for t in range(10):
            ws.cell(row=20, column=fc + t, value=f"Y{t + 1}").font = BOLD
            ws.column_dimensions[self._col(fc + t)].width = 14

        labels = {
            sales_row: "Sales",
            margin_row: "NOPAT Margin",
            nowc_row: "NOWC",
            nola_row: "NOLA",
            nd_row: "Net Debt",
            eq_row: "Book Equity",
            nopat_row: "NOPAT",
            ni_row: "Net Income",
            ae_row: "Abnormal Earnings",
            disc_row: "Discount Factor",
            pv_ae_row: "PV Abnormal Earnings",
        }
        for row, label in labels.items():
            ws.cell(row=row, column=1, value=label)

        ws.cell(row=sum_pv_ae_row, column=1, value="Sum PV Abnormal Earnings")
        ws.cell(row=tv_row, column=1, value="Terminal Value")
        ws.cell(row=tv_pv_row, column=1, value="PV Terminal Value")
        ws.cell(row=iv_row, column=1, value="Intrinsic Value")
        ws.cell(row=shares_row, column=1, value="Diluted Shares")
        ws.cell(row=ivps_row, column=1, value="Intrinsic Value per Share")

        for t in range(10):
            col_i = fc + t
            col = self._col(col_i)
            prev = self._col(col_i - 1) if t > 0 else None
            g = sc["growthVector"][t]
            m = sc["marginVector"][t]
            nowc_ratio = sc["nowcRatioVector"][t]
            nola_ratio = sc["nolaRatioVector"][t]

            if t == 0:
                sales_f = f"='Income Statement'!{anchor_col}{rev_row}*(1+{g})"
            else:
                sales_f = f"={prev}{sales_row}*(1+{g})"
            ws.cell(row=sales_row, column=col_i, value=sales_f).number_format = NUM_FMT

            mc = ws.cell(row=margin_row, column=col_i, value=m)
            mc.font = BLUE
            mc.number_format = PCT_FMT

            if t == 0:
                ws.cell(
                    row=nowc_row,
                    column=col_i,
                    value=f"='Condensed Financials'!{anchor_col}{nowc_hist}",
                ).number_format = NUM_FMT
                ws.cell(row=nola_row, column=col_i, value=nola_hist_formula).number_format = NUM_FMT
                ws.cell(
                    row=nd_row,
                    column=col_i,
                    value=f"='Condensed Financials'!{anchor_col}{nd_hist}",
                ).number_format = NUM_FMT
            else:
                ws.cell(
                    row=nowc_row,
                    column=col_i,
                    value=f"={col}{sales_row}*{nowc_ratio}",
                ).number_format = NUM_FMT
                ws.cell(
                    row=nola_row,
                    column=col_i,
                    value=f"={col}{sales_row}*{nola_ratio}",
                ).number_format = NUM_FMT
                ws.cell(
                    row=nd_row,
                    column=col_i,
                    value=f"=($B$8)*({col}{nowc_row}+{col}{nola_row})",
                ).number_format = NUM_FMT

            ws.cell(
                row=eq_row,
                column=col_i,
                value=f"={col}{nowc_row}+{col}{nola_row}-{col}{nd_row}",
            ).number_format = NUM_FMT
            ws.cell(
                row=nopat_row,
                column=col_i,
                value=f"={col}{sales_row}*{col}{margin_row}",
            ).number_format = NUM_FMT
            if scenario == "Base" and t == 0:
                ws.cell(row=nopat_row, column=col_i).fill = GREEN

            # After-tax CoD in B7 already after-tax — apply once.
            ws.cell(
                row=ni_row,
                column=col_i,
                value=f"={col}{nopat_row}-{col}{nd_row}*$B$7",
            ).number_format = NUM_FMT
            ws.cell(
                row=ae_row,
                column=col_i,
                value=f"={col}{ni_row}-$B$5*{col}{eq_row}",
            ).number_format = NUM_FMT
            if scenario == "Base" and t == 0:
                ws.cell(row=ae_row, column=col_i).fill = GREEN

            ws.cell(
                row=disc_row,
                column=col_i,
                value=f"=1/(1+$B$5)^{t + 1}",
            ).number_format = "0.0000"
            ws.cell(
                row=pv_ae_row,
                column=col_i,
                value=f"={col}{ae_row}*{col}{disc_row}",
            ).number_format = NUM_FMT

        last_fc = self._col(fc + 9)
        first_fc = self._col(fc)
        sum_pv = f"=SUM({first_fc}{pv_ae_row}:{last_fc}{pv_ae_row})"
        ws.cell(row=sum_pv_ae_row, column=fc, value=sum_pv).number_format = NUM_FMT

        tv_formula = f"={last_fc}{ae_row}*(1+$B$9)/($B$5-$B$9)"
        ws.cell(row=tv_row, column=fc + 9, value=tv_formula).number_format = NUM_FMT

        tv_pv_formula = f"={last_fc}{tv_row}*{last_fc}{disc_row}"
        ws.cell(row=tv_pv_row, column=fc, value=tv_pv_formula).number_format = NUM_FMT

        iv_formula = f"={first_fc}{eq_row}+{first_fc}{sum_pv_ae_row}+{first_fc}{tv_pv_row}"
        ws.cell(row=iv_row, column=fc, value=iv_formula).number_format = NUM_FMT

        shares_cell = ws.cell(
            row=shares_row,
            column=fc,
            value=self.assumptions["marketData"]["dilutedShares"],
        )
        shares_cell.font = BLUE
        shares_cell.number_format = NUM_FMT

        ivps_formula = f"={first_fc}{iv_row}/{first_fc}{shares_row}"
        ivps_cell = ws.cell(row=ivps_row, column=fc, value=ivps_formula)
        ivps_cell.number_format = "0.0000"
        if scenario == "Base":
            ivps_cell.fill = GREEN

        self.rowmap[f"model_{scenario}_ivps_row"] = ivps_row
        self.rowmap[f"model_{scenario}_fc_col"] = fc
        self.rowmap[f"model_{scenario}_ae_row"] = ae_row
        self.rowmap[f"model_{scenario}_disc_row"] = disc_row
        self.rowmap[f"model_{scenario}_tv_row"] = tv_row
        self.rowmap[f"model_{scenario}_sales_row"] = sales_row

        if scenario == "Base":
            sales_y1 = str(ws.cell(row=sales_row, column=fc).value)
            nopat_y1 = str(ws.cell(row=nopat_row, column=fc).value)
            ae_y1 = str(ws.cell(row=ae_row, column=fc).value)
            self._register_deferred(
                "model_sales_y1",
                f"Model_{scenario}",
                sales_row,
                fc,
                sales_y1,
                result.sales_y1,
            )
            self._register_deferred(
                "model_nopat_y1",
                f"Model_{scenario}",
                nopat_row,
                fc,
                nopat_y1,
                result.nopat_y1,
            )
            self._register_deferred(
                "model_ae_y1",
                f"Model_{scenario}",
                ae_row,
                fc,
                ae_y1,
                result.abnormal_earnings_y1,
            )
            self._register_deferred(
                "model_tv",
                f"Model_{scenario}",
                tv_pv_row,
                fc,
                tv_pv_formula,
                result.terminal_value_pv,
            )
            self._register_deferred(
                "model_ivps",
                f"Model_{scenario}",
                ivps_row,
                fc,
                ivps_formula,
                result.ivps,
            )

    def _build_scenario_summary(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Scenario_Summary")
        ws["A1"] = f"{self.fin.company_name} — Scenario Summary"
        for j, h in enumerate(["Scenario", "Probability", "IVPS"], start=1):
            ws.cell(row=3, column=j, value=h).font = BOLD
        for i, name in enumerate(("Bear", "Base", "Bull"), start=4):
            sc = self.assumptions["scenarios"][name]
            ivps_row = self.rowmap[f"model_{name}_ivps_row"]
            fc = self.rowmap[f"model_{name}_fc_col"]
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=sc["probability"])
            ws.cell(row=i, column=2).font = BLUE
            ws.cell(row=i, column=2).number_format = PCT_FMT
            ws.cell(row=i, column=3, value=f"='Model_{name}'!{self._col(fc)}{ivps_row}")
        weighted_row = 8
        weighted_col = 5
        weighted_formula = "=SUMPRODUCT(B4:B6,C4:C6)"
        ws.cell(row=weighted_row, column=1, value="Weighted IVPS")
        ws.cell(row=weighted_row, column=weighted_col, value=weighted_formula)
        ws.cell(row=weighted_row, column=weighted_col).fill = GREEN
        w_ivps = weighted_ivps(self._scenario_results, self.assumptions["scenarios"])
        self._register_deferred(
            "scenario_weighted",
            "Scenario_Summary",
            weighted_row,
            weighted_col,
            weighted_formula,
            w_ivps,
        )
