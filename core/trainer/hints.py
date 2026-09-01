"""Progressive hint system for trainer components."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from ..engine.components import TRAINER_COMPONENTS

HINT_FILL = PatternFill("solid", start_color="FFF9E6")
META_SHEET = "_TrainerMeta"


@dataclass
class HintResult:
    component_id: str
    level: int
    max_level: int
    hint_text: str
    related_cells: list[str]
    exhausted: bool


def _component_by_id(component_id: str):
    for c in TRAINER_COMPONENTS:
        if c.id == component_id:
            return c
    raise KeyError(f"Unknown component: {component_id}")


def _find_meta_row(wb, component_id: str) -> int | None:
    if META_SHEET not in wb.sheetnames:
        return None
    ws = wb[META_SHEET]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == component_id:
            return r
    return None


def show_hint(workbook_path: Path, component_id: str) -> HintResult:
    """Reveal the next progressive hint and update metadata."""
    comp = _component_by_id(component_id)
    wb = load_workbook(workbook_path)
    meta_row = _find_meta_row(wb, component_id)
    level = 0
    if meta_row:
        level = int(wb[META_SHEET].cell(row=meta_row, column=7).value or 0)

    if level >= len(comp.hints):
        hint_text = comp.short_hint + " (all detailed hints shown)"
        exhausted = True
    else:
        hint_text = comp.hints[level]
        exhausted = False
        level += 1
        if meta_row:
            wb[META_SHEET].cell(row=meta_row, column=7, value=level)

    if comp.tab in wb.sheetnames:
        from openpyxl.utils import column_index_from_string

        col = column_index_from_string("".join(c for c in comp.cell if c.isalpha()))
        row = int("".join(c for c in comp.cell if c.isdigit()))
        hint_cell = wb[comp.tab].cell(row=row, column=col + 1)
        prefix = f"[Hint {level}/{len(comp.hints)}] "
        hint_cell.value = prefix + hint_text
        hint_cell.fill = HINT_FILL
        hint_cell.font = Font(italic=True, size=9)

    wb.save(workbook_path)
    wb.close()

    return HintResult(
        component_id=component_id,
        level=level,
        max_level=len(comp.hints),
        hint_text=hint_text,
        related_cells=comp.related_cells,
        exhausted=exhausted,
    )


def reveal_answer(workbook_path: Path, component_id: str) -> str:
    """Insert the hidden reference formula into the practice cell."""
    comp = _component_by_id(component_id)
    wb = load_workbook(workbook_path)

    formula = None
    if "_RefFormulas" in wb.sheetnames:
        ws = wb["_RefFormulas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == component_id:
                formula = row[3]
                break

    if not formula:
        ref_path = workbook_path.with_name(
            workbook_path.stem.replace("_Trainer", "") + "_reference.xlsx"
        )
        if ref_path.exists():
            ref_wb = load_workbook(ref_path, data_only=False)
            if comp.tab in ref_wb.sheetnames:
                from openpyxl.utils import column_index_from_string

                col = column_index_from_string("".join(c for c in comp.cell if c.isalpha()))
                row = int("".join(c for c in comp.cell if c.isdigit()))
                formula = ref_wb[comp.tab].cell(row=row, column=col).value
            ref_wb.close()

    if not formula:
        wb.close()
        raise ValueError(f"No reference formula found for {component_id}")

    from openpyxl.utils import column_index_from_string

    col = column_index_from_string("".join(c for c in comp.cell if c.isalpha()))
    row = int("".join(c for c in comp.cell if c.isdigit()))
    cell = wb[comp.tab].cell(row=row, column=col)
    cell.value = formula

    meta_row = _find_meta_row(wb, component_id)
    if meta_row:
        wb[META_SHEET].cell(row=meta_row, column=9, value="revealed")

    if "Trainer" in wb.sheetnames:
        for r in range(5, wb["Trainer"].max_row + 1):
            if wb["Trainer"].cell(row=r, column=3).value == comp.cell:
                if wb["Trainer"].cell(row=r, column=2).value == comp.tab:
                    wb["Trainer"].cell(row=r, column=4, value="revealed")
                    break

    wb.save(workbook_path)
    wb.close()
    return str(formula)
