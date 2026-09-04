"""Training workbook generator — uses runtime SemanticMap as single source of truth."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side

from ..engine.component_catalog import COMPONENT_CATALOG
from ..engine.reference_model import ReferenceModelBuilder
from ..engine.semantic_map import ResolvedComponent, SemanticMap
from ..data.line_identity import validate_financials_identities
from ..ingestion.reconciler import reconcile_financials
from .semantic_io import load_semantic_map, resolve_pair_paths

COMPONENT_MAP_SHEET = "_ComponentMap"
NOTE_AUTHOR = "BAV Trainer"

PRACTICE_FILL = PatternFill("solid", start_color="FFFF00")

FONT_NAME = "Aptos Narrow"
TITLE_FONT = Font(name=FONT_NAME, size=20, bold=True, color="000000")
BODY_FONT = Font(name=FONT_NAME, size=11, color="000000")
BODY_BOLD_FONT = Font(name=FONT_NAME, size=11, bold=True, color="000000")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

_HIDDEN_PREFIX = "_"

TRAINER_INDEX_INSTRUCTION = (
    "Complete each historical schedule left-to-right in dependency order. "
    "Each schedule is one modeling concept repeated across fiscal periods. "
    "Run Check to validate every yellow cell in the workbook. "
    "Open the matching Answer Key for the formula/input and Note hint."
)


class TrainingWorkbookGenerator:
    """Generate a matched Trainer / Answer Key pair from a completed model."""

    def __init__(self, answer_key_path: Path, semantic_map: SemanticMap | None = None):
        self.answer_key_path = answer_key_path
        self.semantic_map = semantic_map or load_semantic_map(answer_key_path)

    def generate(self, trainer_path: Path) -> tuple[Path, Path]:
        """Finalize Answer Key in place, then derive a sanitized Trainer from it."""
        wb = load_workbook(self.answer_key_path)
        self._apply_oshkosh_style(wb)
        self._add_trainer_ui(wb)
        self._decorate_answer_key_practice_cells(wb)
        wb.save(self.answer_key_path)
        wb.close()

        shutil.copy2(self.answer_key_path, trainer_path)

        wb = load_workbook(trainer_path)
        self._blank_trainer_practice_cells(wb)
        self._sanitize_trainer_answer_stores(wb)
        wb.save(trainer_path)
        wb.close()

        # Prefer no Trainer semantic sidecar; Check reads the matching Answer Key.
        return trainer_path, self.answer_key_path

    def _visible_sheets(self, wb):
        return [ws for ws in wb.worksheets if not ws.title.startswith(_HIDDEN_PREFIX)]

    def _apply_oshkosh_style(self, wb) -> None:
        """Apply shared Oshkosh-derived fonts and base styling to visible sheets."""
        for ws in self._visible_sheets(wb):
            ws.sheet_view.showGridLines = False
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    if cell.value is None and (cell.fill is None or cell.fill.fill_type is None):
                        continue
                    is_title = cell.row == 1 and cell.column == 1 and isinstance(cell.value, str)
                    was_bold = bool(cell.font and cell.font.bold)
                    if is_title:
                        cell.font = TITLE_FONT
                    elif was_bold:
                        cell.font = BODY_BOLD_FONT
                    elif cell.value is not None:
                        cell.font = BODY_FONT
                    # White base; practice yellow is applied later
                    if cell.fill and cell.fill.fill_type == "solid":
                        cell.fill = WHITE_FILL

            # Thin borders on header / section / total label rows
            for r in range(1, max_row + 1):
                label = ws.cell(row=r, column=1).value
                if not isinstance(label, str):
                    continue
                upper = label.upper()
                is_section = (
                    upper.isupper() and len(label) > 3 and " " in label
                ) or label.endswith(":")
                is_header = r <= 6 and was_header_row(ws, r)
                is_total = "total" in label.lower() or label.lower().startswith("weighted")
                if is_section or is_header or is_total or (r == 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        if cell.value is not None or is_header:
                            cell.border = THIN_BORDER

    def _decorate_answer_key_practice_cells(self, wb) -> None:
        for comp in self.semantic_map.all_ordered():
            if comp.tab not in wb.sheetnames:
                continue
            ws = wb[comp.tab]
            row, col = _cell_to_rc(comp.cell)
            cell = ws.cell(row=row, column=col)
            # Retain working formula/input; apply bright yellow + legacy Note
            cell.fill = PRACTICE_FILL
            hint = (comp.short_hint or "").strip()
            if not hint and comp.hints:
                hint = str(comp.hints[0]).strip()
            if not hint:
                hint = comp.title
            cell.comment = Comment(hint, NOTE_AUTHOR)

    def _blank_trainer_practice_cells(self, wb) -> None:
        for comp in self.semantic_map.all_ordered():
            if comp.tab not in wb.sheetnames:
                continue
            ws = wb[comp.tab]
            row, col = _cell_to_rc(comp.cell)
            cell = ws.cell(row=row, column=col)
            # Preserve font, border, alignment, protection, number format
            cell.value = None
            cell.fill = PRACTICE_FILL
            cell.comment = None

    def _sanitize_trainer_answer_stores(self, wb) -> None:
        """Remove answer-bearing hidden sheets from the Trainer only."""
        for name in (
            COMPONENT_MAP_SHEET,
            "_RefFormulas",
            "_RefValues",
            "_TrainerMeta",
        ):
            if name in wb.sheetnames:
                del wb[name]

    def _add_trainer_ui(self, wb) -> None:
        if "Trainer" in wb.sheetnames:
            del wb["Trainer"]
        ws = wb.create_sheet("Trainer", 0)
        ws.sheet_view.showGridLines = False
        ws["A1"] = "BAV Excel Trainer"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = TRAINER_INDEX_INSTRUCTION
        ws["A2"].font = BODY_FONT
        headers = ["Order", "Schedule", "Period scope", "Tab", "Practice cells", "Depends on"]
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=j, value=h)
            cell.font = BODY_BOLD_FONT
            cell.border = THIN_BORDER
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 28
        ws.column_dimensions["F"].width = 28

        for i, group in enumerate(group_components_by_family(self.semantic_map), start=5):
            ws.cell(row=i, column=1, value=group["family_order"]).font = BODY_FONT
            ws.cell(row=i, column=2, value=group["title"]).font = BODY_FONT
            ws.cell(row=i, column=3, value=group["period_scope"]).font = BODY_FONT
            ws.cell(row=i, column=4, value=group["tab"]).font = BODY_FONT
            ws.cell(row=i, column=5, value=group["practice_cells"]).font = BODY_FONT
            ws.cell(row=i, column=6, value=group["depends_on"]).font = BODY_FONT


def group_components_by_family(smap: SemanticMap) -> list[dict]:
    """Group concrete ResolvedComponents into conceptual schedule rows."""
    by_family: dict[str, list[ResolvedComponent]] = {}
    for comp in smap.all_ordered():
        by_family.setdefault(comp.family_id or comp.id, []).append(comp)

    family_meta = {f.id: f for f in COMPONENT_CATALOG}
    groups: list[dict] = []
    for family_id, comps in by_family.items():
        comps = sorted(comps, key=lambda c: (c.period_index is None, c.period_index or 0, c.order))
        first = comps[0]
        family = family_meta.get(family_id)
        ends = [c.period_end for c in comps if c.period_end]
        if ends:
            year_span = _period_scope_label(ends, len(comps))
        else:
            year_span = f"{len(comps)} cells"
        dep_ids: list[str] = []
        if family is not None:
            for dep in family.depends_on_current + family.depends_on_previous:
                if dep not in dep_ids:
                    dep_ids.append(dep)
        else:
            for dep in first.depends_on:
                fam = dep.split("__", 1)[0]
                if fam not in dep_ids:
                    dep_ids.append(fam)
        groups.append(
            {
                "family_id": family_id,
                "family_order": first.family_order or first.order,
                "title": first.title,
                "period_scope": year_span,
                "tab": first.tab,
                "practice_cells": _format_practice_cells(comps),
                "depends_on": ", ".join(dep_ids) if dep_ids else "—",
                "count": len(comps),
                "components": comps,
            }
        )
    groups.sort(key=lambda g: g["family_order"])
    return groups


def _period_scope_label(period_ends: list[str], count: int) -> str:
    years = []
    for end in period_ends:
        years.append(end[:4] if len(end) >= 4 else end)
    if len(years) == 1:
        return f"{years[0]} ({count} cells)"
    return f"{years[0]}–{years[-1]} ({count} cells)"


def _format_practice_cells(comps: list[ResolvedComponent]) -> str:
    from openpyxl.utils import column_index_from_string, get_column_letter

    if not comps:
        return "—"
    cells = [c.cell for c in comps]
    if len(cells) == 1:
        return cells[0]

    parsed = []
    for cell in cells:
        col = "".join(ch for ch in cell if ch.isalpha())
        row = int("".join(ch for ch in cell if ch.isdigit()))
        parsed.append((row, column_index_from_string(col), cell))

    rows = {p[0] for p in parsed}
    if len(rows) == 1:
        cols = sorted(p[1] for p in parsed)
        if cols == list(range(cols[0], cols[0] + len(cols))):
            row = next(iter(rows))
            start = f"{get_column_letter(cols[0])}{row}"
            end = f"{get_column_letter(cols[-1])}{row}"
            return f"{start}:{end}"
    return ", ".join(cells)

def was_header_row(ws, row: int) -> bool:
    """Heuristic: row looks like a column-header band (dates / Metric / Scenario)."""
    vals = [ws.cell(row=row, column=c).value for c in range(1, min(6, (ws.max_column or 1) + 1))]
    texts = [v for v in vals if isinstance(v, str)]
    if not texts:
        return False
    markers = ("line item", "metric", "scenario", "probability", "ivps")
    return any(t.lower() in markers or t.lower().startswith("line") for t in texts)


def _cell_to_rc(cell_ref: str) -> tuple[int, int]:
    from openpyxl.utils import column_index_from_string

    col = "".join(c for c in cell_ref if c.isalpha())
    row = int("".join(c for c in cell_ref if c.isdigit()))
    return row, column_index_from_string(col)


def build_training_workbook(
    financials,
    output_path: Path,
    assumptions: dict | None = None,
) -> tuple[Path, Path]:
    """End-to-end: standardized data → Answer Key + Trainer workbook pair.

    Returns ``(trainer_path, answer_key_path)``.
    """
    validate_financials_identities(financials)

    report = reconcile_financials(financials)
    failed = [name for name, ok in report.checksums.items() if ok is False]
    if failed:
        details = "; ".join(failed)
        warnings = "; ".join(report.warnings) if report.warnings else details
        raise ValueError(
            f"Source statement checksum failed ({details}). "
            f"Refuse Trainer / Answer Key build. {warnings}"
        )

    trainer_path, answer_key_path = resolve_pair_paths(output_path)
    builder = ReferenceModelBuilder(financials, assumptions)
    semantic_map = builder.build(answer_key_path)
    TrainingWorkbookGenerator(answer_key_path, semantic_map).generate(trainer_path)
    return trainer_path, answer_key_path
