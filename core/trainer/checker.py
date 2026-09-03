"""Workbook-wide practice-cell validation using the matching Answer Key map."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .semantic_io import answer_key_path_for, load_semantic_map, parse_cell_ref
from .xlsx_fill_patch import CellFillUpdate, apply_fill_updates

BLANK_RGB = "FFFF00"
CORRECT_RGB = "C8E6C9"
INCORRECT_RGB = "FFC7CE"


@dataclass(frozen=True)
class CheckSummary:
    """Non-disclosing aggregate Check result — no formulas, values, or hints."""

    total: int
    correct: int
    incorrect: int
    blank: int


def _normalize_formula(formula: str) -> str:
    return formula.replace(" ", "").replace("'", "").upper()


def _is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _values_match(user_val, expected, tolerance: float) -> bool:
    try:
        u, e = float(user_val), float(expected)
    except (TypeError, ValueError):
        return str(user_val).strip().upper() == str(expected).strip().upper()
    if e == 0:
        return abs(u - e) <= tolerance
    return abs(u - e) / abs(e) <= tolerance or abs(u - e) <= tolerance


def check_workbook(trainer_path: Path) -> CheckSummary:
    """Scan every practice cell in the Trainer; recolor yellow/green/red only.

    Reference semantics come from the matching Answer Key. Never writes answers,
    formulas, expected values, or hints into the Trainer. Fill updates are applied
    via OOXML so formula cached results survive repeated Checks.
    """
    trainer_path = Path(trainer_path)
    answer_key_path = answer_key_path_for(trainer_path)
    if not answer_key_path.exists():
        raise FileNotFoundError(
            f"Answer Key not found for Trainer {trainer_path.name}: "
            f"expected {answer_key_path}"
        )

    smap = load_semantic_map(answer_key_path)
    comps = smap.all_ordered()

    wb = load_workbook(trainer_path, data_only=False)
    wb_cached = load_workbook(trainer_path, data_only=True)

    updates: list[CellFillUpdate] = []
    correct = incorrect = blank = 0
    for comp in comps:
        if comp.tab not in wb.sheetnames:
            incorrect += 1
            continue
        row, col = parse_cell_ref(comp.cell)
        cell = wb[comp.tab].cell(row=row, column=col)
        formula_val = cell.value

        if _is_blank(formula_val):
            updates.append(CellFillUpdate(comp.tab, comp.cell, BLANK_RGB))
            blank += 1
            continue

        if (
            isinstance(formula_val, str)
            and formula_val.startswith("=")
            and _normalize_formula(formula_val) == _normalize_formula(comp.formula)
        ):
            updates.append(CellFillUpdate(comp.tab, comp.cell, CORRECT_RGB))
            correct += 1
            continue

        cached = None
        if comp.tab in wb_cached.sheetnames:
            cached = wb_cached[comp.tab].cell(row=row, column=col).value

        if cached is not None and _values_match(cached, comp.expected_value, comp.tolerance):
            updates.append(CellFillUpdate(comp.tab, comp.cell, CORRECT_RGB))
            correct += 1
        else:
            updates.append(CellFillUpdate(comp.tab, comp.cell, INCORRECT_RGB))
            incorrect += 1

    wb.close()
    wb_cached.close()
    apply_fill_updates(trainer_path, updates)

    return CheckSummary(
        total=len(comps),
        correct=correct,
        incorrect=incorrect,
        blank=blank,
    )
