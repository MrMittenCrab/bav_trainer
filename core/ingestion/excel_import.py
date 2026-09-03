"""Excel / Bloomberg / Wind export import adapter."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from ..data.interface import (
    DocumentManifest,
    DocumentType,
    FinancialPeriod,
    LineItem,
    StandardizedFinancials,
)
from ..data.schema import normalize_label
from .base import BaseIngestionAdapter

TAB_MAP = {
    "income statement": "income_statement",
    "income_statement": "income_statement",
    "is": "income_statement",
    "balance sheet": "balance_sheet",
    "balance_sheet": "balance_sheet",
    "bs": "balance_sheet",
    "cash flow": "cash_flow",
    "cash flow statement": "cash_flow",
    "cash_flow": "cash_flow",
    "cf": "cash_flow",
}


def _parse_header(cell_value) -> date | None:
    if isinstance(cell_value, date):
        return cell_value
    if cell_value is None:
        return None
    text = str(cell_value).strip()
    for fmt in ("%Y-%m-%d", "%b %d, %Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    if text.isdigit() and len(text) == 4:
        return date(int(text), 12, 31)
    return None


def _header_layout(header: tuple) -> tuple[int | None, int, int]:
    """Return (concept_col, label_col, first_date_col)."""
    if not header:
        return None, 0, 1
    h0 = str(header[0] or "").strip().lower()
    h1 = str(header[1] or "").strip().lower() if len(header) > 1 else ""
    label_aliases = {"line item", "label", "line_item", "lineitem"}
    if h0 == "concept" and h1 in label_aliases:
        return 0, 1, 2
    # Legacy: optional "Line Item" header in col 0, dates from col 1
    return None, 0, 1


class ExcelExportAdapter(BaseIngestionAdapter):
    """Import IS/BS/CF from Excel exports (Bloomberg, Wind, manual transcription)."""

    jurisdiction = "HK"

    def ingest(self, manifest: list[DocumentManifest]) -> StandardizedFinancials:
        if len(manifest) != 1:
            raise ValueError("Excel adapter accepts one workbook per ingest call")
        path = Path(manifest[0].path)
        meta = manifest[0]
        wb = load_workbook(path, data_only=True, read_only=True)

        ticker = path.stem.split("_")[0].upper()
        company = ticker
        currency = "HKD"
        units = "HKD in Millions"

        stmts: dict[str, list[LineItem]] = {
            "income_statement": [],
            "balance_sheet": [],
            "cash_flow": [],
        }
        periods: list[FinancialPeriod] = []

        for sheet in wb.worksheets:
            key = TAB_MAP.get(sheet.title.strip().lower())
            if not key:
                continue
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            header = rows[0]
            concept_col, label_col, date_start = _header_layout(header)
            col_dates: list[tuple[int, date]] = []
            for ci, hv in enumerate(header[date_start:], start=date_start):
                pd = _parse_header(hv)
                if pd:
                    col_dates.append((ci, pd))
            if not periods and col_dates:
                periods = [FinancialPeriod(end_date=d) for _, d in col_dates]
            items: list[LineItem] = []
            for row in rows[1:]:
                if not row or label_col >= len(row) or not row[label_col]:
                    continue
                label = normalize_label(str(row[label_col]))
                concept = ""
                if concept_col is not None and concept_col < len(row) and row[concept_col]:
                    concept = normalize_label(str(row[concept_col]))
                values = {}
                for ci, pd in col_dates:
                    val = row[ci] if ci < len(row) else None
                    if isinstance(val, (int, float)):
                        values[pd] = float(val)
                    else:
                        values[pd] = None
                items.append(
                    LineItem(
                        label=label,
                        values=values,
                        concept=concept,
                        source_doc=str(path),
                    )
                )
            stmts[key] = items

        wb.close()
        doc_type = meta.doc_type
        jurisdiction = "HK"
        if doc_type == DocumentType.BLOOMBERG_EXPORT:
            units = meta.notes or units
        elif doc_type == DocumentType.WIND_EXPORT:
            units = meta.notes or units

        return StandardizedFinancials(
            ticker=ticker,
            company_name=company,
            currency=currency,
            units=units,
            jurisdiction=jurisdiction,
            stock_code=meta.notes if meta.doc_type == DocumentType.OTHER else "",
            periods=periods,
            income_statement=stmts["income_statement"],
            balance_sheet=stmts["balance_sheet"],
            cash_flow=stmts["cash_flow"],
            provenance=[{"source": str(path), "type": meta.doc_type.value}],
        )
