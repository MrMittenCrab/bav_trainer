"""Formula-integrity tests for the semantic Answer Key (Step 2)."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
import pytest

from core.data.interface import DocumentManifest, DocumentType
from core.engine.component_catalog import COMPONENT_CATALOG
from core.ingestion.manual_hk import HKManualDocumentAdapter
from core.model.classification import BALANCE_SHEET_CATEGORIES
from core.model.financial_math import _val, compute_anchor
from core.model.line_resolver import resolve_line
from core.model.ri_engine import run_scenario
from core.trainer.semantic_io import load_semantic_map, parse_cell_ref
from core.trainer.workbook import build_training_workbook

ROOT = Path(__file__).resolve().parents[2]
DEMO_JSON = ROOT / "example" / "DEMO_HK_Standardized.json"


def _ingest_demo():
    adapter = HKManualDocumentAdapter()
    return adapter.ingest([DocumentManifest(path=str(DEMO_JSON), doc_type=DocumentType.OTHER)])


def _build_pair(tmp_path, assumptions=None):
    data = _ingest_demo()
    return build_training_workbook(data, tmp_path / "DEMO_HK_Trainer.xlsx", assumptions)


def test_normal_v1_build_does_not_call_run_scenario(tmp_path, monkeypatch):
    import core.engine.reference_model as rm

    def fail(*args, **kwargs):
        raise AssertionError("forecast engine executed in historical-only v1")

    monkeypatch.setattr(rm, "run_scenario", fail)

    data = _ingest_demo()
    trainer, answer = build_training_workbook(
        data,
        tmp_path / "DEMO_HK_Trainer.xlsx",
    )

    assert trainer.exists()
    assert answer.exists()


def test_normal_v1_build_requires_no_forecast_assumptions(tmp_path):
    data = _ingest_demo()
    trainer, answer = build_training_workbook(
        data,
        tmp_path / "DEMO_HK_Trainer.xlsx",
        assumptions={"classificationOverrides": {}},
    )
    assert trainer.exists()
    assert answer.exists()
    sidecar = answer.with_suffix(".assumptions.json")
    if sidecar.exists():
        payload = json.loads(sidecar.read_text(encoding="utf-8"))
        assert "scenarios" not in payload
        assert "growthVector" not in json.dumps(payload)
        assert "marginVector" not in json.dumps(payload)
        assert "terminalGrowth" not in json.dumps(payload)
        assert "beta" not in json.dumps(payload)
        market = payload.get("marketData") or {}
        assert "dilutedShares" not in market


def test_anchor_exposes_tax_and_interest_expected_values():
    data = _ingest_demo()
    periods = data.fiscal_years() or data.period_dates()
    anchor = compute_anchor(data, periods)
    last = periods[-1]

    pretax = resolve_line(data.income_statement, "pretax_income", required=True).item
    tax = resolve_line(data.income_statement, "tax_expense", required=True).item
    int_exp = resolve_line(data.income_statement, "interest_expense", required=False).item
    int_inc = resolve_line(data.income_statement, "interest_income", required=False).item
    ni = resolve_line(data.income_statement, "net_income", required=True).item

    pretax_v = _val(pretax, last)
    tax_v = _val(tax, last)
    expected_etr = (-tax_v / pretax_v) if pretax_v else 0.0
    expected_net_int = -(_val(int_exp, last) + _val(int_inc, last))

    assert anchor.effective_tax_rate == pytest.approx(expected_etr)
    assert anchor.net_interest == pytest.approx(expected_net_int)
    assert anchor.net_interest_after_tax == pytest.approx(
        anchor.net_interest * (1 - anchor.effective_tax_rate)
    )
    assert anchor.nopat == pytest.approx(_val(ni, last) + anchor.net_interest_after_tax)


HISTORICAL_REFORMULATION_IDS = (
    "effective_tax_rate_fy",
    "net_interest_fy",
    "net_interest_after_tax_fy",
    "nopat_fy",
    "owca_agg",
    "owcl_agg",
    "nowc_agg",
    "olta_agg",
    "oltl_agg",
    "nola_agg",
    "noa_agg",
    "financial_assets_agg",
    "financial_liabilities_agg",
    "net_debt",
    "equity_reformulated_fy",
)

AGGREGATE_SUMIF_IDS = (
    "owca_agg",
    "owcl_agg",
    "olta_agg",
    "oltl_agg",
    "financial_assets_agg",
    "financial_liabilities_agg",
)

DUPONT_IDS = (
    "rnoa",
    "after_tax_cod",
    "spread",
    "flev",
    "roe_decomp",
    "actual_roe",
)


def test_cli_assumptions_propagate(tmp_path):
    """CLI-supplied historical classification overrides must affect reformulation."""
    data = _ingest_demo()
    _, answer_default = build_training_workbook(data, tmp_path / "Default_Trainer.xlsx")
    smap_default = load_semantic_map(answer_default)

    # Move Cash into operating WC so NOA changes without breaking BS integrity.
    assumptions = {
        "classificationOverrides": {
            "label:Cash and cash equivalents": "Operating Working Capital Asset",
        }
    }
    from core.__main__ import main

    out = tmp_path / "Assumed_Trainer.xlsx"
    assumptions_path = tmp_path / "custom.assumptions.json"
    assumptions_path.write_text(json.dumps(assumptions), encoding="utf-8")
    assert main(["build", str(DEMO_JSON), "-o", str(out), "-a", str(assumptions_path)]) == 0

    answer = tmp_path / "Assumed_Answer_Key.xlsx"
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    smap = load_semantic_map(answer)
    found = False
    for r in range(1, 40):
        if ws.cell(row=r, column=1).value == "Cash and cash equivalents":
            assert ws.cell(row=r, column=2).value == "Operating Working Capital Asset"
            found = True
            break
    assert found
    assert smap.get("nowc_agg").expected_value != smap_default.get("nowc_agg").expected_value
    wb.close()


def test_nopat_formula_adds_after_tax_net_interest(tmp_path):
    _, answer = _build_pair(tmp_path)
    smap = load_semantic_map(answer)
    nopat = smap.get("nopat_fy")
    formula = nopat.formula.replace(" ", "")
    assert formula.startswith("=")
    assert "+" in formula
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    assert "Net Interest" in labels
    assert "Net Interest After Tax" in labels
    assert "NOPAT" in labels
    niat_r = labels["Net Interest After Tax"]
    ni_r = labels["Net Income"]
    nopat_r = labels["NOPAT"]
    row, col = parse_cell_ref(nopat.cell)
    cell_formula = str(ws.cell(row=row, column=col).value)
    assert row == nopat_r
    assert str(niat_r) in cell_formula.replace(" ", "")
    assert str(ni_r) in cell_formula.replace(" ", "")
    data = _ingest_demo()
    periods = data.fiscal_years() or data.period_dates()
    anchor = compute_anchor(data, periods)
    assert nopat.expected_value == anchor.nopat
    from core.model.line_resolver import resolve_line
    from core.model.financial_math import _val

    ni_item = resolve_line(data.income_statement, "net_income", required=True).item
    assert abs(float(nopat.expected_value) - float(_val(ni_item, periods[-1]))) > 1.0
    wb.close()


def test_condensed_aggregates_are_on_sheet_sumif(tmp_path):
    _, answer = _build_pair(tmp_path)
    smap = load_semantic_map(answer)
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    for cid in ("nowc_agg", "noa_agg", "net_debt"):
        formula = smap.get(cid).formula
        assert "Balance Sheet" not in formula
    # Category SUMIFs live on the detail aggregate rows that feed NOWC / NOA / Net Debt.
    for label in (
        "Operating Working Capital Assets",
        "Operating Long-Term Assets",
        "Financial Liabilities",
    ):
        f = str(ws.cell(row=labels[label], column=2).value)
        assert "SUMIF" in f and "$B$" in f
    wb.close()


def test_dupont_cod_uses_niat_not_nopat(tmp_path):
    _, answer = _build_pair(tmp_path)
    wb = load_workbook(answer, data_only=False)
    ws = wb["ALT DuPont"]
    cod_row = None
    for r in range(1, 20):
        if ws.cell(row=r, column=1).value == "After-tax CoD":
            cod_row = r
            break
    assert cod_row is not None
    formula = str(ws.cell(row=cod_row, column=5).value)
    assert "NOPAT" not in formula
    condensed = wb["Condensed Financials"]
    niat_row = next(
        r for r in range(1, condensed.max_row + 1)
        if condensed.cell(row=r, column=1).value == "Net Interest After Tax"
    )
    nopat_row = next(
        r for r in range(1, condensed.max_row + 1)
        if condensed.cell(row=r, column=1).value == "NOPAT"
    )
    assert str(niat_row) in formula
    assert str(nopat_row) not in formula.split("/")[0]
    wb.close()


def test_dupont_uses_condensed_equity_not_bs_row_seven(tmp_path):
    _, answer = _build_pair(tmp_path)
    wb = load_workbook(answer, data_only=False)
    ws = wb["ALT DuPont"]
    condensed = wb["Condensed Financials"]
    eq_row = next(
        r for r in range(1, condensed.max_row + 1)
        if condensed.cell(row=r, column=1).value == "Equity (NOA - Net Debt)"
    )
    flev_row = next(r for r in range(1, 20) if ws.cell(row=r, column=1).value == "FLEV")
    actual_row = next(r for r in range(1, 20) if ws.cell(row=r, column=1).value == "Actual ROE")
    flev_f = str(ws.cell(row=flev_row, column=5).value)
    actual_f = str(ws.cell(row=actual_row, column=5).value)
    assert "Balance Sheet" not in flev_f
    assert "Balance Sheet" not in actual_f
    assert "Condensed Financials" in flev_f and str(eq_row) in flev_f
    assert "Condensed Financials" in actual_f and str(eq_row) in actual_f
    wb.close()


def test_ten_year_forecast_chain_populated(tmp_path):
    """Internal deferred-forecast path still builds a full 10-year chain."""
    data = _ingest_demo()
    from core.engine.reference_model import ReferenceModelBuilder

    answer = tmp_path / "Deferred_Answer_Key.xlsx"
    ReferenceModelBuilder(data, include_deferred_forecast=True).build(answer)
    wb = load_workbook(answer, data_only=False)
    for scenario in ("Bear", "Base", "Bull"):
        ws = wb[f"Model_{scenario}"]
        # Locate Sales row and first forecast col from Y1 header
        sales_row = next(r for r in range(1, 50) if ws.cell(row=r, column=1).value == "Sales")
        ae_row = next(
            r for r in range(1, 50) if ws.cell(row=r, column=1).value == "Abnormal Earnings"
        )
        disc_row = next(
            r for r in range(1, 50) if ws.cell(row=r, column=1).value == "Discount Factor"
        )
        tv_row = next(r for r in range(1, 50) if ws.cell(row=r, column=1).value == "Terminal Value")
        # First forecast column: first Y1 header
        fc = next(
            c for c in range(2, 30)
            if ws.cell(row=20, column=c).value == "Y1"
        )
        for t in range(10):
            assert ws.cell(row=sales_row, column=fc + t).value not in (None, "")
            assert ws.cell(row=ae_row, column=fc + t).value not in (None, "")
            assert ws.cell(row=disc_row, column=fc + t).value not in (None, "")
            for label in (
                "NOPAT Margin",
                "NOWC",
                "NOLA",
                "Net Debt",
                "Book Equity",
                "NOPAT",
                "Net Income",
                "PV Abnormal Earnings",
            ):
                row = next(r for r in range(1, 50) if ws.cell(row=r, column=1).value == label)
                assert ws.cell(row=row, column=fc + t).value not in (None, ""), f"{scenario} {label} Y{t+1}"
        # Year-10 TV populated; PV TV references year-10 cells
        tv_cell = ws.cell(row=tv_row, column=fc + 9).value
        assert tv_cell not in (None, "")
        assert str(ae_row) in str(tv_cell)
        pv_tv_row = next(
            r for r in range(1, 50) if ws.cell(row=r, column=1).value == "PV Terminal Value"
        )
        pv_tv = str(ws.cell(row=pv_tv_row, column=fc).value)
        assert str(tv_row) in pv_tv
        assert str(disc_row) in pv_tv
        # No double-tax of after-tax CoD
        assert ws["B8"].value != "=B7*(1-B6)"
        assert ws["A7"].value == "After-tax CoD"
    wb.close()


def test_semantic_formulas_have_no_blank_required_refs(tmp_path):
    import re

    _, answer = _build_pair(tmp_path)
    smap = load_semantic_map(answer)
    wb = load_workbook(answer, data_only=False)
    assert len(smap.all_ordered()) == len(COMPONENT_CATALOG)

    def _labels(ws):
        return {ws.cell(row=r, column=1).value: r for r in range(1, (ws.max_row or 1) + 1)}

    def _a1_refs(formula: str) -> list[tuple[str | None, str]]:
        """Extract optional sheet + A1 refs from a formula (test-only helper)."""
        refs = []
        for m in re.finditer(
            r"(?:'([^']+)'!)?(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)",
            formula,
        ):
            refs.append((m.group(1), m.group(2).replace("$", "")))
        return refs

    def _cell_populated(sheet: str, a1: str) -> bool:
        if ":" in a1:
            # range — check both ends
            start, end = a1.split(":")
            return _cell_populated(sheet, start) and _cell_populated(sheet, end)
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        col_letter, row = coordinate_from_string(a1)
        val = wb[sheet].cell(row=row, column=column_index_from_string(col_letter)).value
        return val not in (None, "")

    for comp in smap.all_ordered():
        assert comp.expected_value is not None
        assert isinstance(comp.formula, str) and comp.formula.startswith("=")
        row, col = parse_cell_ref(comp.cell)
        cell_val = wb[comp.tab].cell(row=row, column=col).value
        assert isinstance(cell_val, str) and cell_val.startswith("=")
        assert cell_val == comp.formula
        for sheet, a1 in _a1_refs(comp.formula):
            target = sheet or comp.tab
            if target.startswith("_"):
                continue
            assert _cell_populated(target, a1.split(":")[0] if ":" in a1 else a1), (
                f"{comp.id} references blank {target}!{a1}"
            )

    condensed = wb["Condensed Financials"]
    cl = _labels(condensed)
    nopat = smap.get("nopat_fy")
    assert str(cl["Net Income"]) in nopat.formula
    assert str(cl["Net Interest After Tax"]) in nopat.formula
    assert condensed.cell(row=cl["Net Income"], column=2).value not in (None, "")
    assert condensed.cell(row=cl["Net Interest After Tax"], column=2).value not in (None, "")

    assert "Condensed Financials" in smap.get("rnoa").formula
    assert smap.get("roe_decomp").formula.startswith("=")
    # Spread is local RNOA − CoD on the DuPont sheet
    spread_f = smap.get("spread").formula
    assert spread_f.startswith("=")
    assert "-" in spread_f

    # Deferred forecast tabs are placeholders, not live practice surfaces.
    for name in ("Model_Bear", "Model_Base", "Model_Bull", "Scenario_Summary"):
        ws = wb[name]
        assert ws.sheet_state == "hidden"
        assert ws["A1"].value == "Deferred from historical-only v1"
    wb.close()


def test_python_expected_values_use_corrected_cod(tmp_path):
    data = _ingest_demo()
    periods = data.fiscal_years() or data.period_dates()
    anchor = compute_anchor(data, periods)
    sc = {
        "costOfEquity": 0.10,
        "terminalGrowth": 0.03,
        "taxRate": 0.165,
        "growthVector": [0.1] * 10,
        "marginVector": [0.15] * 10,
        "nowcRatioVector": [0.05] * 10,
        "nolaRatioVector": [0.5] * 10,
    }
    r_ok = run_scenario(sc, anchor, shares=1000.0)
    taxed_again = anchor.hist_avg_after_tax_cod * (1 - 0.165)
    r_bug = run_scenario(sc, anchor, shares=1000.0, hist_avg_after_tax_cod=taxed_again)
    assert r_ok.abnormal_earnings_y1 != r_bug.abnormal_earnings_y1

    # Internal deferred path still builds live model tabs for legacy isolation.
    from core.engine.reference_model import ReferenceModelBuilder

    answer = tmp_path / "Deferred_Answer_Key.xlsx"
    builder = ReferenceModelBuilder(data, include_deferred_forecast=True)
    builder.build(answer)
    smap = load_semantic_map(answer)
    wb = load_workbook(answer, data_only=False)
    ws = wb["Model_Base"]
    ni_row = next(r for r in range(1, 50) if ws.cell(row=r, column=1).value == "Net Income")
    fc = next(c for c in range(2, 30) if ws.cell(row=20, column=c).value == "Y1")
    ni_f = str(ws.cell(row=ni_row, column=fc).value)
    assert "$B$7" in ni_f
    assert smap.get("model_ae_y1").expected_value is not None
    wb.close()


def test_pair_behavior_still_holds(tmp_path):
    trainer, answer = _build_pair(tmp_path)
    assert trainer.exists() and answer.exists()
    smap = load_semantic_map(answer)
    wb_t = load_workbook(trainer, data_only=False)
    wb_a = load_workbook(answer, data_only=False)
    for comp in smap.all_ordered():
        row, col = parse_cell_ref(comp.cell)
        ct = wb_t[comp.tab].cell(row=row, column=col)
        ca = wb_a[comp.tab].cell(row=row, column=col)
        assert ct.value is None
        assert ct.comment is None
        assert isinstance(ca.value, str) and ca.value.startswith("=")
        assert ca.comment is not None and ca.comment.text
    wb_t.close()
    wb_a.close()


# --- Step 2B source-resolution fixtures ---

def _synth_periods():
    from datetime import date
    from core.data.interface import FinancialPeriod

    return [
        FinancialPeriod(end_date=date(2024, 12, 31), label="FY2024"),
        FinancialPeriod(end_date=date(2025, 12, 31), label="FY2025"),
    ]


def _li(label, v1, v2, concept=""):
    from datetime import date
    from core.data.interface import LineItem

    return LineItem(
        label=label,
        values={date(2024, 12, 31): v1, date(2025, 12, 31): v2},
        concept=concept,
    )


def _base_fin(**overrides):
    from core.data.interface import StandardizedFinancials

    periods = _synth_periods()
    is_items = [
        _li("Revenue", 1000, 1100),
        _li("Finance costs", -40, -50),
        _li("Finance income", 5, 6),
        _li("Profit before tax", 200, 220),
        _li("Income tax expense", -30, -33),
        _li("Profit for the year", 170, 187),
    ]
    bs_items = [
        _li("Cash and cash equivalents", 100, 110),
        _li("Trade receivables", 80, 90),
        _li("Property, plant and equipment", 400, 420),
        _li("Trade payables", 50, 55),
        _li("Bank borrowings", 200, 210),
        _li("Total equity", 330, 355),
    ]
    cf_items = [_li("Net cash from operating activities", 50, 60)]
    kwargs = dict(
        ticker="SYN",
        company_name="Synthetic Co",
        currency="HKD",
        units="HKD mn",
        jurisdiction="HK",
        periods=periods,
        income_statement=is_items,
        balance_sheet=bs_items,
        cash_flow=cf_items,
    )
    kwargs.update(overrides)
    return StandardizedFinancials(**kwargs)


def test_missing_interest_income_still_populates_net_interest_chain(tmp_path):
    is_items = [
        _li("Revenue", 1000, 1100),
        _li("Finance costs", -40, -50),
        _li("Profit before tax", 200, 220),
        _li("Income tax expense", -30, -33),
        _li("Profit for the year", 170, 187),
    ]
    fin = _base_fin(income_statement=is_items)
    periods = [p.end_date for p in fin.periods]
    anchor = compute_anchor(fin, periods)
    assert anchor.nopat != 0
    _, answer = build_training_workbook(fin, tmp_path / "MissInc_Trainer.xlsx")
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    assert "Interest Income" not in labels
    assert "Interest Expense" in labels
    for name in ("Net Interest", "Net Interest After Tax", "NOPAT"):
        row = labels[name]
        assert ws.cell(row=row, column=2).value not in (None, "")
    wb.close()


def test_missing_interest_expense_income_only_case(tmp_path):
    is_items = [
        _li("Revenue", 1000, 1100),
        _li("Finance income", 5, 6),
        _li("Profit before tax", 200, 220),
        _li("Income tax expense", -30, -33),
        _li("Profit for the year", 170, 187),
    ]
    fin = _base_fin(income_statement=is_items)
    periods = [p.end_date for p in fin.periods]
    anchor = compute_anchor(fin, periods)
    assert anchor.nopat != 0
    _, answer = build_training_workbook(fin, tmp_path / "MissExp_Trainer.xlsx")
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    assert "Interest Expense" not in labels
    assert "Interest Income" in labels
    for name in ("Net Interest", "Net Interest After Tax", "NOPAT"):
        assert ws.cell(row=labels[name], column=2).value not in (None, "")
    wb.close()


def test_equity_alias_builds_and_matches_python(tmp_path):
    bs = [
        _li("Cash and cash equivalents", 100, 110),
        _li("Trade receivables", 80, 90),
        _li("Property, plant and equipment", 400, 420),
        _li("Trade payables", 50, 55),
        _li("Bank borrowings", 200, 210),
        _li("Equity attributable to owners of the Company", 330, 355),
    ]
    fin = _base_fin(balance_sheet=bs)
    periods = [p.end_date for p in fin.periods]
    anchor = compute_anchor(fin, periods)
    assert anchor.equity == 355.0
    _, answer = build_training_workbook(fin, tmp_path / "EqAlias_Trainer.xlsx")
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    implied = str(ws.cell(row=labels["Equity (NOA - Net Debt)"], column=3).value)
    reported = str(ws.cell(row=labels["Reported Equity"], column=3).value)
    assert "Balance Sheet" not in implied
    assert str(labels["NOA"]) in implied and str(labels["Net Debt"]) in implied
    assert "Balance Sheet" in reported
    wb.close()


def test_equity_absent_fallback_noa_minus_net_debt(tmp_path):
    bs = [
        _li("Cash and cash equivalents", 100, 110),
        _li("Trade receivables", 80, 90),
        _li("Property, plant and equipment", 400, 420),
        _li("Trade payables", 50, 55),
        _li("Bank borrowings", 200, 210),
    ]
    fin = _base_fin(balance_sheet=bs)
    periods = [p.end_date for p in fin.periods]
    anchor = compute_anchor(fin, periods)
    assert abs(anchor.equity - (anchor.noa - anchor.net_debt)) < 1e-9
    _, answer = build_training_workbook(fin, tmp_path / "EqFallback_Trainer.xlsx")
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    f = str(ws.cell(row=labels["Equity (NOA - Net Debt)"], column=2).value)
    assert "Balance Sheet" not in f
    assert str(labels["NOA"]) in f and str(labels["Net Debt"]) in f
    check_f = str(ws.cell(row=labels["CHECK"], column=2).value)
    assert "UNVERIFIED" in check_f
    wb.close()


def test_build_rejects_failed_source_checksum(tmp_path):
    data = _ingest_demo()
    # Break CF roll-up while leaving other statements intact.
    for item in data.cash_flow:
        if "operating activities" in item.label.lower():
            for pd in list(item.values):
                item.values[pd] = float(item.values[pd]) + 999.0
    import pytest

    with pytest.raises(ValueError, match="checksum"):
        build_training_workbook(data, tmp_path / "BrokenCF_Trainer.xlsx")


def test_build_rejects_reformulation_gap(tmp_path):
    from datetime import date
    from core.data.interface import FinancialPeriod, StandardizedFinancials
    from core.model.classification import ReformulationIntegrityError
    import pytest

    p1, p2 = date(2024, 12, 31), date(2025, 12, 31)
    fin = StandardizedFinancials(
        ticker="GAP",
        company_name="Gap Co",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=[
            FinancialPeriod(end_date=p1, label="FY2024"),
            FinancialPeriod(end_date=p2, label="FY2025"),
        ],
        income_statement=[
            _li("Revenue", 100, 110),
            _li("Profit before tax", 20, 22),
            _li("Income tax expense", -3, -3),
            _li("Profit for the year", 17, 19),
        ],
        balance_sheet=[
            _li("Cash and cash equivalents", 40, 40),
            _li("Trade receivables", 30, 30),
            _li("Total assets", 100, 100),
            _li("Trade payables", 20, 20),
            _li("Bank borrowings", 30, 30),
            _li("Total liabilities", 80, 80),
            _li("Share capital and reserves", 20, 20),
            _li("Total equity", 20, 20),
        ],
        cash_flow=[
            _li("Net cash from operating activities", 10, 10),
            _li("Net cash used in investing activities", -4, -4),
            _li("Net cash from financing activities", -1, -1),
            _li("Net change in cash and cash equivalents", 5, 5),
        ],
    )
    with pytest.raises(ReformulationIntegrityError):
        build_training_workbook(fin, tmp_path / "Gap_Trainer.xlsx")


def test_condensed_has_live_reconciliation_rows(tmp_path):
    _, answer = _build_pair(tmp_path)
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    labels = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    required = [
        "Operating Working Capital Assets",
        "Operating Working Capital Liabilities",
        "NOWC",
        "Operating Long-Term Assets",
        "Operating Long-Term Liabilities",
        "NOLA",
        "NOA",
        "Financial Assets",
        "Financial Liabilities",
        "Net Debt",
        "Equity (NOA - Net Debt)",
        "Reported Equity",
        "Total Capital",
        "CHECK",
    ]
    for name in required:
        assert name in labels, name
    check_f = str(ws.cell(row=labels["CHECK"], column=2).value)
    assert check_f.startswith("=")
    assert "OK" in check_f and "CHECK" in check_f
    wb.close()


def test_classification_table_uses_shared_decisions(tmp_path):
    from core.model.classification import BALANCE_SHEET_CATEGORIES
    from core.engine.reference_model import ReferenceModelBuilder
    from core.model.financial_math import compute_anchor
    from core.data.interface import FinancialPeriod, StandardizedFinancials
    from datetime import date

    data = _ingest_demo()
    b = ReferenceModelBuilder(data)
    assumptions = b.assumptions
    assumptions["classificationOverrides"] = {
        "Goodwill": "Operating Long-Term Asset",
    }

    _, answer = build_training_workbook(
        data, tmp_path / "ClassDemo_Trainer.xlsx", assumptions
    )

    periods = data.fiscal_years() or data.period_dates()
    anchor = compute_anchor(
        data,
        periods,
        classification_overrides={"Goodwill": "Operating Long-Term Asset"},
    )
    wb = load_workbook(answer, data_only=False)
    ws = wb["Condensed Financials"]
    start = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Line Item":
            start = r + 1
            break
    assert start is not None
    notes_col = 3 + len(periods)
    by_label = {}
    r = start
    while ws.cell(row=r, column=1).value and ws.cell(row=r, column=2).value:
        label = ws.cell(row=r, column=1).value
        if label in ("CONDENSED INCOME STATEMENT", "CONDENSED BALANCE SHEET"):
            break
        by_label[label] = r
        r += 1
    for idx, decision in anchor.reformulation.decisions.items():
        label = data.balance_sheet[idx].label
        row = by_label[label]
        assert ws.cell(row=row, column=2).value == decision.category
        if decision.overridden:
            note = str(ws.cell(row=row, column=notes_col).value or "")
            assert "Override" in note
    cats = set()
    for dv in ws.data_validations.dataValidation:
        if dv.formula1:
            cats |= {c.strip() for c in dv.formula1.strip('"').split(",")}
    assert cats == set(BALANCE_SHEET_CATEGORIES)
    assert not any(c.startswith("Ambiguous") for c in cats)
    wb.close()

    p1, p2 = date(2024, 12, 31), date(2025, 12, 31)
    fin = StandardizedFinancials(
        ticker="AMB",
        company_name="Amb Co",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=[
            FinancialPeriod(end_date=p1, label="FY2024"),
            FinancialPeriod(end_date=p2, label="FY2025"),
        ],
        income_statement=[
            _li("Revenue", 100, 110),
            _li("Profit before tax", 20, 22),
            _li("Income tax expense", -3, -3),
            _li("Profit for the year", 17, 19),
        ],
        balance_sheet=[
            _li("Cash and cash equivalents", 40, 40),
            _li("Trade receivables", 30, 30),
            _li("Property, plant and equipment", 30, 30),
            _li("Total assets", 100, 100),
            _li("Trade payables", 20, 20),
            _li("Operating lease liabilities", 30, 30),
            _li("Bank borrowings", 30, 30),
            _li("Total liabilities", 80, 80),
            _li("Share capital and reserves", 20, 20),
            _li("Total equity", 20, 20),
        ],
        cash_flow=[
            _li("Net cash from operating activities", 10, 10),
            _li("Net cash used in investing activities", -4, -4),
            _li("Net cash from financing activities", -1, -1),
            _li("Net change in cash and cash equivalents", 5, 5),
        ],
    )
    _, answer2 = build_training_workbook(fin, tmp_path / "Amb_Trainer.xlsx")
    wb2 = load_workbook(answer2, data_only=False)
    ws2 = wb2["Condensed Financials"]
    lease_row = next(
        r
        for r in range(1, ws2.max_row + 1)
        if ws2.cell(row=r, column=1).value == "Operating lease liabilities"
    )
    note = str(ws2.cell(row=lease_row, column=5).value or "")
    assert "⚠ Review" in note
    wb2.close()


def test_duPont_uses_implied_equity(tmp_path):
    _, answer = _build_pair(tmp_path)
    wb = load_workbook(answer, data_only=False)
    condensed = wb["Condensed Financials"]
    ws = wb["ALT DuPont"]
    implied_row = next(
        r
        for r in range(1, condensed.max_row + 1)
        if condensed.cell(row=r, column=1).value == "Equity (NOA - Net Debt)"
    )
    reported_row = next(
        r
        for r in range(1, condensed.max_row + 1)
        if condensed.cell(row=r, column=1).value == "Reported Equity"
    )
    flev_row = next(r for r in range(1, 20) if ws.cell(row=r, column=1).value == "FLEV")
    actual_row = next(r for r in range(1, 20) if ws.cell(row=r, column=1).value == "Actual ROE")
    flev_f = str(ws.cell(row=flev_row, column=5).value)
    actual_f = str(ws.cell(row=actual_row, column=5).value)
    assert str(implied_row) in flev_f
    assert str(implied_row) in actual_f
    assert str(reported_row) not in flev_f
    assert str(reported_row) not in actual_f
    wb.close()


def test_historical_reformulation_formulas_match_answer_key_cells(tmp_path):
    trainer_path, answer = _build_pair(tmp_path)
    smap = load_semantic_map(answer)
    wb = load_workbook(answer, data_only=False)
    for cid in HISTORICAL_REFORMULATION_IDS:
        comp = smap.get(cid)
        assert comp.formula.startswith("=")
        assert comp.expected_value is not None
        row, col = parse_cell_ref(comp.cell)
        assert wb[comp.tab].cell(row=row, column=col).value == comp.formula
        if cid in AGGREGATE_SUMIF_IDS:
            assert "SUMIF" in comp.formula
    wb.close()

    # Classification choices remain populated identically in Trainer vs Answer Key.
    wb_t = load_workbook(trainer_path, data_only=False)
    wb_a = load_workbook(answer, data_only=False)
    class_start = None
    for r in range(1, 40):
        if wb_a["Condensed Financials"].cell(row=r, column=1).value == "Line Item":
            class_start = r + 1
            break
    assert class_start is not None
    for r in range(class_start, class_start + 50):
        label = wb_a["Condensed Financials"].cell(row=r, column=1).value
        cat = wb_a["Condensed Financials"].cell(row=r, column=2).value
        if not isinstance(label, str) or not label.strip():
            break
        if cat is None:
            continue
        assert cat in BALANCE_SHEET_CATEGORIES
        assert (
            wb_t["Condensed Financials"].cell(row=r, column=2).value
            == wb_a["Condensed Financials"].cell(row=r, column=2).value
        )
    wb_t.close()
    wb_a.close()


def test_dupont_chain_registers_all_six_latest_comparable_formulas(tmp_path):
    _, answer = _build_pair(tmp_path)
    data = _ingest_demo()
    periods = data.fiscal_years() or data.period_dates()
    anchor = compute_anchor(data, periods)
    j_last = len(periods) - 1
    smap = load_semantic_map(answer)
    wb = load_workbook(answer, data_only=False)

    expected_map = {
        "rnoa": anchor.dupont["RNOA"][j_last],
        "after_tax_cod": anchor.dupont["After-tax CoD"][j_last],
        "spread": anchor.dupont["Spread"][j_last],
        "flev": anchor.dupont["FLEV"][j_last],
        "roe_decomp": anchor.dupont["ROE (decomposed)"][j_last],
        "actual_roe": anchor.dupont["Actual ROE"][j_last],
    }
    for cid in DUPONT_IDS:
        comp = smap.get(cid)
        assert comp.formula.startswith("=")
        row, col = parse_cell_ref(comp.cell)
        assert wb[comp.tab].cell(row=row, column=col).value == comp.formula
        assert comp.expected_value == pytest.approx(expected_map[cid])
    wb.close()

    by_id = {c.id: c for c in COMPONENT_CATALOG}
    for spec in COMPONENT_CATALOG:
        for dep in spec.depends_on:
            assert by_id[dep].order < spec.order
