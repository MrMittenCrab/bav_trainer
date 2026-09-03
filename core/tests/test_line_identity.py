"""Step 4 — concept-aware statement line identity."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.data.interface import (
    DocumentManifest,
    DocumentType,
    FinancialPeriod,
    LineItem,
    StandardizedFinancials,
)
from core.data.line_identity import (
    AmbiguousStatementIdentityError,
    line_identity,
    validate_statement_identities,
)
from core.ingestion.excel_import import ExcelExportAdapter
from core.ingestion.reconciler import merge_documents, _merge_line_items
from core.model.classification import (
    classify_balance_sheet_line,
    reformulate_balance_sheet,
)
from core.trainer.workbook import build_training_workbook

P1 = date(2024, 12, 31)
P2 = date(2025, 12, 31)


def _li(label: str, v1: float, v2: float, concept: str = "") -> LineItem:
    return LineItem(label=label, values={P1: v1, P2: v2}, concept=concept)


def _periods() -> list[FinancialPeriod]:
    return [
        FinancialPeriod(end_date=P1, label="FY2024"),
        FinancialPeriod(end_date=P2, label="FY2025"),
    ]


def test_line_identity_label_is_case_insensitive_but_concept_is_exact():
    a = _li("Goodwill", 10, 12, "Goodwill")
    b = _li("  GOODWILL  ", 10, 12, "Goodwill")
    assert line_identity(a) == line_identity(b)

    nbsp = _li("Goodwill\u00a0", 10, 12, "Goodwill")
    assert line_identity(a) == line_identity(nbsp)

    # Concept IDs remain exact identifiers; case-only concept changes are not
    # silently assumed to be the same taxonomy concept.
    c = _li("Goodwill", 10, 12, "goodwill")
    assert line_identity(a) != line_identity(c)


def test_conceptless_duplicate_labels_differing_only_by_case_fail():
    items = [
        _li("Goodwill", 10, 12),
        _li("GOODWILL", 11, 13),
    ]
    with pytest.raises(AmbiguousStatementIdentityError):
        validate_statement_identities(items, "balance_sheet")


def test_same_concept_and_case_variant_label_merge_as_one_identity():
    existing = [_li("Goodwill", 10, 12, "Goodwill")]
    incoming = [_li("GOODWILL", 10.1, 12.1, "Goodwill")]

    merged, _ = _merge_line_items(existing, incoming, "restatement")

    assert len(merged) == 1
    assert merged[0].label == "Goodwill"  # original display label preserved
    assert merged[0].values[P2] == 12.1


def _deferred_pair() -> list[LineItem]:
    return [
        _li("Deferred income taxes", 40, 45, "DeferredIncomeTaxAssetsNet"),
        _li("Deferred income taxes", 25, 28, "DeferredIncomeTaxLiabilitiesNet"),
    ]


def test_distinct_concepts_same_label_survive_merge():
    """Regression: label-keyed merge must not collapse asset/liability deferred tax."""
    existing = _deferred_pair()
    incoming = [
        _li("Deferred income taxes", 50, 55, "DeferredIncomeTaxAssetsNet"),
        _li("Deferred income taxes", 30, 32, "DeferredIncomeTaxLiabilitiesNet"),
    ]
    merged, conflicts = _merge_line_items(existing, incoming, "restatement")
    assert len(merged) == 2
    by_concept = {line_identity(i).concept: i for i in merged}
    assert set(by_concept) == {
        "DeferredIncomeTaxAssetsNet",
        "DeferredIncomeTaxLiabilitiesNet",
    }
    # Large period deltas are conflict-logged and keep the existing value.
    assert by_concept["DeferredIncomeTaxAssetsNet"].values[P2] == 45
    assert by_concept["DeferredIncomeTaxLiabilitiesNet"].values[P2] == 28
    assert len(conflicts) >= 1
    assert all("concept" in c and "label" in c for c in conflicts)
    assert {c["concept"] for c in conflicts} >= {
        "DeferredIncomeTaxAssetsNet",
        "DeferredIncomeTaxLiabilitiesNet",
    }


def test_correct_classification_by_concept():
    asset = _li("Deferred income taxes", 40, 45, "DeferredIncomeTaxAssetsNet")
    liab = _li("Deferred income taxes", 25, 28, "DeferredIncomeTaxLiabilitiesNet")
    da = classify_balance_sheet_line(asset)
    dl = classify_balance_sheet_line(liab)
    assert "Asset" in da.category
    assert "Liability" in dl.category
    assert da.category != dl.category


def test_concept_specific_overrides():
    items = _deferred_pair()
    fin = StandardizedFinancials(
        ticker="DT",
        company_name="DT Co",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        balance_sheet=items
        + [
            _li("Cash and cash equivalents", 100, 100),
            _li("Total assets", 185, 193),
            _li("Bank borrowings", 100, 100),
            _li("Total liabilities", 125, 128),
            _li("Share capital and reserves", 60, 65),
            _li("Total equity", 60, 65),
        ],
        income_statement=[
            _li("Revenue", 100, 110),
            _li("Profit before tax", 20, 22),
            _li("Income tax expense", -3, -3),
            _li("Profit for the year", 17, 19),
        ],
        cash_flow=[
            _li("Net cash from operating activities", 10, 10),
            _li("Net cash used in investing activities", -4, -4),
            _li("Net cash from financing activities", -1, -1),
            _li("Net change in cash and cash equivalents", 5, 5),
        ],
    )
    overrides = {
        "concept:DeferredIncomeTaxAssetsNet": "Financial Asset",
        "concept:DeferredIncomeTaxLiabilitiesNet": "Financial Liability",
    }
    reform = reformulate_balance_sheet(fin, [P1, P2], overrides=overrides)
    assert reform.decisions[0].category == "Financial Asset"
    assert reform.decisions[0].overridden is True
    assert reform.decisions[1].category == "Financial Liability"
    assert reform.decisions[1].overridden is True


def test_ambiguous_label_override_rejected():
    items = _deferred_pair()
    fin = StandardizedFinancials(
        ticker="DT",
        company_name="DT Co",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        balance_sheet=items,
    )
    with pytest.raises(Exception) as exc:
        reformulate_balance_sheet(
            fin,
            [P1, P2],
            overrides={"label:Deferred income taxes": "Exclude"},
        )
    msg = str(exc.value).lower()
    assert "deferred income taxes" in msg or "ambiguous" in msg


def test_legacy_unique_label_override_still_works():
    item = _li("Goodwill", 10, 12)
    fin = StandardizedFinancials(
        ticker="GW",
        company_name="GW Co",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        balance_sheet=[item],
    )
    reform = reformulate_balance_sheet(
        fin, [P1, P2], overrides={"Goodwill": "Operating Long-Term Asset"}
    )
    assert reform.decisions[0].category == "Operating Long-Term Asset"
    assert reform.decisions[0].overridden is True


def test_legacy_unique_label_override_is_case_insensitive():
    item = _li("Goodwill", 10, 12)
    fin = StandardizedFinancials(
        ticker="GW",
        company_name="GW Co",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        balance_sheet=[item],
    )
    reform = reformulate_balance_sheet(
        fin, [P1, P2], overrides={"goodwill": "Operating Long-Term Asset"}
    )
    assert reform.decisions[0].category == "Operating Long-Term Asset"
    assert reform.decisions[0].overridden is True

    reform2 = reformulate_balance_sheet(
        fin, [P1, P2], overrides={"label:GOODWILL": "Exclude"}
    )
    assert reform2.decisions[0].category == "Exclude"
    assert reform2.decisions[0].overridden is True


def test_duplicate_identity_inside_existing_fails_before_merge():
    existing = [
        _li("Deferred income taxes", 40, 45, "DeferredIncomeTaxAssetsNet"),
        _li("Deferred income taxes", 41, 46, "DeferredIncomeTaxAssetsNet"),
    ]
    with pytest.raises(AmbiguousStatementIdentityError):
        _merge_line_items(existing, [], "source")


def test_duplicate_identity_inside_incoming_fails_before_merge():
    incoming = [
        _li("Deferred income taxes", 40, 45, "DeferredIncomeTaxAssetsNet"),
        _li("Deferred income taxes", 41, 46, "DeferredIncomeTaxAssetsNet"),
    ]
    with pytest.raises(AmbiguousStatementIdentityError):
        _merge_line_items([], incoming, "source")


def test_unconcepted_duplicate_labels_fail(tmp_path):
    fin = StandardizedFinancials(
        ticker="AMB",
        company_name="Amb",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        income_statement=[_li("Revenue", 1, 1), _li("Profit for the year", 1, 1)],
        balance_sheet=[
            _li("Deferred income taxes", 10, 10),
            _li("Deferred income taxes", 5, 5),
            _li("Cash and cash equivalents", 100, 100),
            _li("Total assets", 115, 115),
            _li("Total liabilities", 15, 15),
            _li("Total equity", 100, 100),
        ],
        cash_flow=[
            _li("Net cash from operating activities", 1, 1),
            _li("Net cash used in investing activities", 0, 0),
            _li("Net cash from financing activities", 0, 0),
            _li("Net change in cash and cash equivalents", 1, 1),
        ],
    )
    with pytest.raises(AmbiguousStatementIdentityError):
        validate_statement_identities(fin.balance_sheet, "balance_sheet")
    with pytest.raises(AmbiguousStatementIdentityError):
        build_training_workbook(fin, tmp_path / "Dup_Trainer.xlsx")


def test_concepted_plus_conceptless_duplicate_fails():
    items = [
        _li("Deferred income taxes", 10, 10, "DeferredIncomeTaxAssetsNet"),
        _li("Deferred income taxes", 5, 5),
    ]
    with pytest.raises(AmbiguousStatementIdentityError):
        validate_statement_identities(items, "balance_sheet")


def test_same_identity_across_documents_merges_restates():
    base = StandardizedFinancials(
        ticker="DT",
        company_name="DT",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        balance_sheet=[
            _li("Deferred income taxes", 40, 45, "DeferredIncomeTaxAssetsNet"),
        ],
    )
    # Small restatement (<2%) updates in place; large restatement logs conflict.
    small = StandardizedFinancials(
        ticker="DT",
        company_name="DT",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        balance_sheet=[
            _li("Deferred income taxes", 40.5, 45.5, "DeferredIncomeTaxAssetsNet"),
        ],
    )
    report_small = merge_documents(base, small, "minor restatement")
    assert len(base.balance_sheet) == 1
    assert base.balance_sheet[0].values[P2] == 45.5

    large = StandardizedFinancials(
        ticker="DT",
        company_name="DT",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        balance_sheet=[
            _li("Deferred income taxes", 50, 55, "DeferredIncomeTaxAssetsNet"),
        ],
    )
    report = merge_documents(base, large, "FY2025 restatement")
    assert len(base.balance_sheet) == 1
    assert base.balance_sheet[0].values[P2] == 45.5  # large conflict keeps prior
    assert any(c.get("concept") == "DeferredIncomeTaxAssetsNet" for c in report.conflicts)


def test_rowmap_preserves_both_duplicate_labels(tmp_path):
    fin = StandardizedFinancials(
        ticker="DT",
        company_name="DT Co",
        currency="HKD",
        units="mn",
        jurisdiction="HK",
        periods=_periods(),
        income_statement=[
            _li("Revenue", 100, 110),
            _li("Profit before tax", 20, 22),
            _li("Income tax expense", -3, -3),
            _li("Profit for the year", 17, 19),
        ],
        balance_sheet=[
            _li("Cash and cash equivalents", 100, 100),
            _li("Deferred income taxes", 40, 45, "DeferredIncomeTaxAssetsNet"),
            _li("Deferred income taxes", 25, 28, "DeferredIncomeTaxLiabilitiesNet"),
            _li("Total assets", 140, 145),
            _li("Bank borrowings", 80, 80),
            _li("Total liabilities", 105, 108),
            _li("Share capital and reserves", 35, 37),
            _li("Total equity", 35, 37),
        ],
        cash_flow=[
            _li("Net cash from operating activities", 10, 10),
            _li("Net cash used in investing activities", -4, -4),
            _li("Net cash from financing activities", -1, -1),
            _li("Net change in cash and cash equivalents", 5, 5),
        ],
    )
    _, answer = build_training_workbook(fin, tmp_path / "DT_Trainer.xlsx")
    import json

    rowmap = json.loads((tmp_path / "rowmap.json").read_text(encoding="utf-8"))
    asset_item = next(
        i for i in fin.balance_sheet if i.concept == "DeferredIncomeTaxAssetsNet"
    )
    liab_item = next(
        i for i in fin.balance_sheet if i.concept == "DeferredIncomeTaxLiabilitiesNet"
    )
    asset_key = f"Balance Sheet!{line_identity(asset_item).key()}"
    liab_key = f"Balance Sheet!{line_identity(liab_item).key()}"
    assert asset_key in rowmap
    assert liab_key in rowmap
    assert rowmap[asset_key] != rowmap[liab_key]
    assert "label=deferred income taxes" in asset_key
    assert "label=deferred income taxes" in liab_key

    from openpyxl import load_workbook

    wb = load_workbook(answer, data_only=False)
    ws = wb["Balance Sheet"]
    labels = [
        ws.cell(row=r, column=1).value
        for r in range(7, (ws.max_row or 7) + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert labels.count("Deferred income taxes") == 2
    wb.close()


def test_excel_optional_concept_column(tmp_path):
    legacy = tmp_path / "legacy.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.append(["Line Item", P1, P2])
    ws.append(["Cash and cash equivalents", 10, 12])
    wb.save(legacy)
    wb.close()

    adapter = ExcelExportAdapter()
    data = adapter.ingest(
        [DocumentManifest(path=str(legacy), doc_type=DocumentType.EXCEL_EXPORT)]
    )
    assert len(data.balance_sheet) == 1
    assert data.balance_sheet[0].concept == ""
    assert data.balance_sheet[0].label == "Cash and cash equivalents"

    with_concept = tmp_path / "with_concept.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.append(["Concept", "Line Item", P1, P2])
    ws.append(["DeferredIncomeTaxAssetsNet", "Deferred income taxes", 40, 45])
    ws.append(["DeferredIncomeTaxLiabilitiesNet", "Deferred income taxes", 25, 28])
    wb.save(with_concept)
    wb.close()

    data2 = adapter.ingest(
        [DocumentManifest(path=str(with_concept), doc_type=DocumentType.EXCEL_EXPORT)]
    )
    assert len(data2.balance_sheet) == 2
    concepts = {i.concept for i in data2.balance_sheet}
    assert concepts == {
        "DeferredIncomeTaxAssetsNet",
        "DeferredIncomeTaxLiabilitiesNet",
    }
    assert all(i.label == "Deferred income taxes" for i in data2.balance_sheet)


def test_cli_ingest_preserves_concept_through_json_round_trip(tmp_path):
    """cmd_ingest -o must export concept so reload keeps deferred-tax identity."""
    from core.__main__ import main

    src = tmp_path / "concept_source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.append(["Concept", "Line Item", P1, P2])
    ws.append(["DeferredIncomeTaxAssetsNet", "Deferred income taxes", 40, 45])
    ws.append(["DeferredIncomeTaxLiabilitiesNet", "Deferred income taxes", 25, 28])
    # Minimal companion statements so ingest can succeed structurally
    for title in ("Income Statement", "Cash Flow"):
        w = wb.create_sheet(title)
        w.append(["Line Item", P1, P2])
        w.append(["Revenue" if title == "Income Statement" else "Net change in cash", 1, 2])
    wb.save(src)
    wb.close()

    out_json = tmp_path / "roundtrip.json"
    rc = main(["ingest", str(src), "-o", str(out_json)])
    assert rc in (0, 1)  # checksum may fail on minimal fixture; concept must still export
    assert out_json.exists()

    payload = __import__("json").loads(out_json.read_text(encoding="utf-8"))
    bs_rows = payload["balance_sheet"]
    by_label = [r for r in bs_rows if r["label"] == "Deferred income taxes"]
    assert len(by_label) == 2
    concepts = {r["concept"] for r in by_label}
    assert concepts == {
        "DeferredIncomeTaxAssetsNet",
        "DeferredIncomeTaxLiabilitiesNet",
    }
    # Explicit empty concept for conceptless rows
    for row in payload["income_statement"] + payload["cash_flow"]:
        assert "concept" in row
        assert isinstance(row["concept"], str)

    from core.ingestion.manual_hk import HKManualDocumentAdapter

    adapter = HKManualDocumentAdapter()
    reloaded = adapter.ingest(
        [DocumentManifest(path=str(out_json), doc_type=DocumentType.OTHER)]
    )
    assert len(reloaded.balance_sheet) >= 2
    reloaded_concepts = {
        i.concept for i in reloaded.balance_sheet if i.label == "Deferred income taxes"
    }
    assert reloaded_concepts == {
        "DeferredIncomeTaxAssetsNet",
        "DeferredIncomeTaxLiabilitiesNet",
    }
    # Still separately identifiable
    ids = {
        line_identity(i)
        for i in reloaded.balance_sheet
        if i.label == "Deferred income taxes"
    }
    assert len(ids) == 2

