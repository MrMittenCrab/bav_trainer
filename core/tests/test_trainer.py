"""Tests for BAV Excel Trainer — Trainer / Answer Key separation + workbook-wide Check."""

import importlib.util
from pathlib import Path

import pytest
from openpyxl import load_workbook

from core.data.interface import DocumentManifest, DocumentType
from core.engine.component_catalog import COMPONENT_CATALOG
from core.ingestion.manual_hk import HKManualDocumentAdapter
from core.trainer.checker import check_workbook
from core.trainer.semantic_io import (
    answer_key_path_for,
    load_semantic_map,
    parse_cell_ref,
    resolve_pair_paths,
)
from core.trainer.workbook import FONT_NAME, TRAINER_INDEX_INSTRUCTION, build_training_workbook

ROOT = Path(__file__).resolve().parents[2]
DEMO_JSON = ROOT / "example" / "DEMO_HK_Standardized.json"


def _fill_rgb(cell) -> str:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return ""
    color = fill.fgColor.rgb or fill.start_color.rgb or ""
    return str(color).upper().lstrip("0")[-6:] if color else ""


def _ingest_demo():
    adapter = HKManualDocumentAdapter()
    return adapter.ingest([DocumentManifest(path=str(DEMO_JSON), doc_type=DocumentType.OTHER)])


def _build_pair(tmp_path):
    data = _ingest_demo()
    return build_training_workbook(data, tmp_path / "DEMO_HK_Trainer.xlsx")


def test_legacy_components_module_removed():
    assert importlib.util.find_spec("core.engine.components") is None


def test_no_trainer_components_symbols_in_repo():
    for path in ROOT.rglob("*.py"):
        if ".git" in path.parts or "__pycache__" in path.parts or "tests" in path.parts:
            continue
        text = path.read_text(encoding="utf-8")
        assert "TRAINER_COMPONENTS" not in text
        assert "TrainerComponent" not in text
        assert "from core.engine.components" not in text
        assert "from .engine.components" not in text


def test_catalog_has_no_coordinates():
    for spec in COMPONENT_CATALOG:
        assert spec.semantic_key
        assert spec.tab_template or spec.category
        d = spec.__dict__
        assert "cell" not in d
        assert "tab" not in d or spec.tab_template


def test_ingest_demo_json():
    adapter = HKManualDocumentAdapter()
    data = adapter.ingest([DocumentManifest(path=str(DEMO_JSON), doc_type=DocumentType.OTHER)])
    assert data.ticker == "DEMO"
    assert len(data.periods) == 5
    report = adapter.reconcile(data)
    assert "income_statement" in report.checksums


def test_resolve_pair_paths_appends_and_preserves_trainer_stem(tmp_path):
    trainer, answer = resolve_pair_paths(tmp_path / "DEMO_HK_Trainer.xlsx")
    assert trainer.name == "DEMO_HK_Trainer.xlsx"
    assert answer.name == "DEMO_HK_Answer_Key.xlsx"

    trainer2, answer2 = resolve_pair_paths(tmp_path / "Acme.xlsx")
    assert trainer2.name == "Acme_Trainer.xlsx"
    assert answer2.name == "Acme_Answer_Key.xlsx"


def test_build_paired_trainer_and_answer_key(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)

    assert trainer_path.exists()
    assert answer_key_path.exists()
    assert trainer_path.name == "DEMO_HK_Trainer.xlsx"
    assert answer_key_path.name == "DEMO_HK_Answer_Key.xlsx"
    assert not (tmp_path / "DEMO_HK_Trainer_reference.xlsx").exists()
    assert not list(tmp_path.glob("*_reference.xlsx"))

    smap = load_semantic_map(answer_key_path)
    assert len(smap.all_ordered()) == len(COMPONENT_CATALOG)
    for spec in COMPONENT_CATALOG:
        comp = smap.get(spec.id)
        assert comp.formula.startswith("=")
        assert comp.expected_value is not None
        assert comp.tab and comp.cell


def test_trainer_has_no_answer_bearing_hidden_sheets(tmp_path):
    trainer_path, _ = _build_pair(tmp_path)
    wb_t = load_workbook(trainer_path, data_only=False)
    assert "_RefFormulas" not in wb_t.sheetnames
    assert "_RefValues" not in wb_t.sheetnames
    assert "_TrainerMeta" not in wb_t.sheetnames
    assert "_ComponentMap" not in wb_t.sheetnames
    wb_t.close()


def test_trainer_has_no_answer_sidecars(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    assert not trainer_path.with_suffix(".component_map.json").exists()
    assert not trainer_path.with_suffix(".trainer.json").exists()
    assert answer_key_path.with_suffix(".component_map.json").exists()


def test_trainer_and_answer_key_practice_contract(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    wb_t = load_workbook(trainer_path, data_only=False)
    wb_a = load_workbook(answer_key_path, data_only=False)
    for comp in smap.all_ordered():
        row, col = parse_cell_ref(comp.cell)
        tc = wb_t[comp.tab].cell(row=row, column=col)
        ac = wb_a[comp.tab].cell(row=row, column=col)
        assert tc.value is None
        assert tc.comment is None
        assert _fill_rgb(tc) == "FFFF00"
        assert isinstance(ac.value, str) and ac.value.startswith("=")
        assert ac.value == comp.formula
        assert ac.comment is not None and ac.comment.text.strip()
        assert _fill_rgb(ac) == "FFFF00"
    wb_t.close()
    wb_a.close()


def test_trainer_does_not_leak_answer_metadata(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    wb_t = load_workbook(trainer_path, data_only=False)

    forbidden: list[str] = []
    for comp in smap.all_ordered():
        hint = (comp.short_hint or "").strip()
        if hint:
            forbidden.append(hint)
        for h in comp.hints:
            if h and h.strip():
                forbidden.append(h.strip())
        if comp.formula:
            forbidden.append(comp.formula)

    # Scan answer-bearing locations only: hidden sheets + comments + Trainer index extras
    for name in wb_t.sheetnames:
        if not name.startswith("_") and name != "Trainer":
            continue
        ws = wb_t[name]
        for row in ws.iter_rows(max_row=ws.max_row or 1, max_col=ws.max_column or 1):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                for needle in forbidden:
                    assert needle not in val, f"leak in {name}: {needle!r}"
                if cell.comment is not None:
                    for needle in forbidden:
                        assert needle not in (cell.comment.text or "")

    for comp in smap.all_ordered():
        row, col = parse_cell_ref(comp.cell)
        cell = wb_t[comp.tab].cell(row=row, column=col)
        assert cell.comment is None
        assert cell.value is None

    wb_t.close()


def test_trainer_index_instruction_and_columns(tmp_path):
    trainer_path, _ = _build_pair(tmp_path)
    wb = load_workbook(trainer_path, data_only=False)
    ws = wb["Trainer"]
    assert ws["A2"].value == TRAINER_INDEX_INSTRUCTION
    headers = [ws.cell(row=4, column=c).value for c in range(1, 6)]
    assert headers == ["Order", "Component", "Tab", "Cell", "Depends on"]
    assert "Status" not in headers
    text = " ".join(
        str(ws.cell(row=r, column=c).value or "")
        for r in range(1, (ws.max_row or 1) + 1)
        for c in range(1, 6)
    )
    assert "HintActive" not in text
    assert "RevealActive" not in text
    assert "TrainerMacros" not in text
    wb.close()


def test_answer_key_practice_cells_formula_yellow_legacy_notes(tmp_path):
    _, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    wb = load_workbook(answer_key_path, data_only=False)
    for comp in smap.all_ordered():
        row, col = parse_cell_ref(comp.cell)
        cell = wb[comp.tab].cell(row=row, column=col)
        assert isinstance(cell.value, str) and cell.value.startswith("=")
        assert cell.value == comp.formula
        assert _fill_rgb(cell) == "FFFF00"
        assert cell.comment is not None
        expected_hint = (comp.short_hint or "").strip() or (
            comp.hints[0] if comp.hints else comp.title
        )
        assert cell.comment.text == expected_hint
        assert cell.comment.author == "BAV Trainer"
    wb.close()


def test_pair_style_and_structure_parity(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    wb_t = load_workbook(trainer_path, data_only=False)
    wb_a = load_workbook(answer_key_path, data_only=False)

    visible_t = [s for s in wb_t.sheetnames if not s.startswith("_")]
    visible_a = [s for s in wb_a.sheetnames if not s.startswith("_")]
    assert visible_t == visible_a

    for name in visible_t:
        ws_t, ws_a = wb_t[name], wb_a[name]
        assert ws_t.freeze_panes == ws_a.freeze_panes
        assert bool(ws_t.sheet_view.showGridLines) == bool(ws_a.sheet_view.showGridLines)
        assert list(ws_t.merged_cells.ranges) == list(ws_a.merged_cells.ranges)
        for letter in ("A", "B", "C"):
            assert ws_t.column_dimensions[letter].width == ws_a.column_dimensions[letter].width
        for r in range(1, min(8, (ws_t.max_row or 1) + 1)):
            assert ws_t.row_dimensions[r].height == ws_a.row_dimensions[r].height

    for comp in smap.all_ordered():
        row, col = parse_cell_ref(comp.cell)
        ct = wb_t[comp.tab].cell(row=row, column=col)
        ca = wb_a[comp.tab].cell(row=row, column=col)
        assert ct.font.name == ca.font.name
        assert ct.font.size == ca.font.size
        assert ct.font.bold == ca.font.bold
        assert ct.border.left.style == ca.border.left.style
        assert ct.alignment.horizontal == ca.alignment.horizontal
        assert ct.alignment.vertical == ca.alignment.vertical
        assert ct.number_format == ca.number_format
        assert bool(ct.protection.locked) == bool(ca.protection.locked)
        assert _fill_rgb(ct) == _fill_rgb(ca) == "FFFF00"

    wb_t.close()
    wb_a.close()


def test_oshkosh_font_conventions(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    for path in (trainer_path, answer_key_path):
        wb = load_workbook(path, data_only=False)
        for name in wb.sheetnames:
            if name.startswith("_"):
                continue
            ws = wb[name]
            title = ws["A1"]
            assert title.font.name == FONT_NAME
            assert title.font.size == 20
            assert title.font.bold is True
            body = ws.cell(row=6, column=1)
            if body.value is not None:
                assert body.font.name == FONT_NAME
                assert body.font.size == 11
        wb.close()


def test_no_adjacent_hint_cells_on_generation(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    for path in (trainer_path, answer_key_path):
        wb = load_workbook(path, data_only=False)
        for comp in smap.all_ordered():
            row, col = parse_cell_ref(comp.cell)
            adj = wb[comp.tab].cell(row=row, column=col + 1)
            assert adj.value != comp.short_hint
        wb.close()


def test_cli_list_catalog():
    from core.__main__ import main

    assert main(["list"]) == 0


def test_check_scans_all_practice_cells_and_colors_three_states(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    comps = smap.all_ordered()

    wb = load_workbook(trainer_path, data_only=False)
    c0 = comps[0]
    r, c = parse_cell_ref(c0.cell)
    wb[c0.tab].cell(r, c).value = c0.formula

    c1 = comps[1]
    r, c = parse_cell_ref(c1.cell)
    wb[c1.tab].cell(r, c).value = "=1+1"
    wb.save(trainer_path)
    wb.close()

    summary = check_workbook(trainer_path)
    assert summary.total == len(comps)
    assert summary.correct == 1
    assert summary.incorrect == 1
    assert summary.blank == len(comps) - 2

    wb = load_workbook(trainer_path, data_only=False)
    r, c = parse_cell_ref(c0.cell)
    assert _fill_rgb(wb[c0.tab].cell(r, c)) == "C8E6C9"
    r, c = parse_cell_ref(c1.cell)
    assert _fill_rgb(wb[c1.tab].cell(r, c)) == "FFC7CE"
    for comp in comps[2:]:
        r, c = parse_cell_ref(comp.cell)
        assert _fill_rgb(wb[comp.tab].cell(r, c)) == "FFFF00"
    wb.close()


def test_recheck_refreshes_colors_from_current_contents(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    comp = smap.all_ordered()[0]
    row, col = parse_cell_ref(comp.cell)

    wb = load_workbook(trainer_path, data_only=False)
    wb[comp.tab].cell(row, col).value = "=1+1"
    wb.save(trainer_path)
    wb.close()

    summary = check_workbook(trainer_path)
    assert summary.incorrect >= 1
    wb = load_workbook(trainer_path, data_only=False)
    assert _fill_rgb(wb[comp.tab].cell(row, col)) == "FFC7CE"

    wb[comp.tab].cell(row, col).value = comp.formula
    wb.save(trainer_path)
    wb.close()
    summary = check_workbook(trainer_path)
    assert summary.correct >= 1
    wb = load_workbook(trainer_path, data_only=False)
    assert _fill_rgb(wb[comp.tab].cell(row, col)) == "C8E6C9"

    wb[comp.tab].cell(row, col).value = None
    wb.save(trainer_path)
    wb.close()
    summary = check_workbook(trainer_path)
    assert summary.blank == summary.total
    wb = load_workbook(trainer_path, data_only=False)
    assert _fill_rgb(wb[comp.tab].cell(row, col)) == "FFFF00"
    wb.close()


def test_check_does_not_change_practice_contents_or_add_notes(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    comps = smap.all_ordered()

    wb = load_workbook(trainer_path, data_only=False)
    c0 = comps[0]
    r, c = parse_cell_ref(c0.cell)
    wb[c0.tab].cell(r, c).value = c0.formula
    c1 = comps[1]
    r, c = parse_cell_ref(c1.cell)
    wb[c1.tab].cell(r, c).value = "=1+1"
    wb.save(trainer_path)
    wb.close()

    before = {}
    wb = load_workbook(trainer_path, data_only=False)
    for comp in comps:
        r, c = parse_cell_ref(comp.cell)
        before[comp.id] = wb[comp.tab].cell(r, c).value
    wb.close()

    check_workbook(trainer_path)

    wb = load_workbook(trainer_path, data_only=False)
    for comp in comps:
        r, c = parse_cell_ref(comp.cell)
        cell = wb[comp.tab].cell(r, c)
        assert cell.value == before[comp.id]
        assert cell.comment is None
    wb.close()


def test_check_requires_matching_answer_key(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    answer_key_path.unlink()
    answer_key_path.with_suffix(".component_map.json").unlink(missing_ok=True)
    with pytest.raises(FileNotFoundError, match="Answer Key"):
        check_workbook(trainer_path)


def test_cli_check_is_workbook_wide_and_hint_reveal_are_removed(tmp_path, capsys):
    from core.__main__ import main
    import io
    from contextlib import redirect_stdout, redirect_stderr
    from core import __main__ as cli

    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)

    rc = main(["check", "--workbook", str(trainer_path)])
    captured = capsys.readouterr()
    assert rc == 0
    assert f"Checked {len(smap.all_ordered())} practice cells" in captured.out
    assert "blank" in captured.out

    buf = io.StringIO()
    with redirect_stdout(buf), redirect_stderr(buf):
        try:
            cli.main(["--help"])
        except SystemExit:
            pass
    help_text = buf.getvalue()
    assert "check" in help_text
    # Subcommand names must not include hint/reveal
    assert "\n  hint " not in help_text and not help_text.strip().startswith("hint ")
    assert "  hint\n" not in help_text and "{hint}" not in help_text
    assert "  reveal" not in help_text and "{reveal}" not in help_text
    assert "hint" not in help_text.split("{")[1].split("}")[0] if "{" in help_text else True

    with pytest.raises(SystemExit):
        cli.main(["check", "--workbook", str(trainer_path), "--component", "nopat_fy"])


def test_cli_check_output_does_not_disclose_answers(tmp_path, capsys):
    from core.__main__ import main

    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)

    wb = load_workbook(trainer_path, data_only=False)
    comps = smap.all_ordered()
    r, c = parse_cell_ref(comps[0].cell)
    wb[comps[0].tab].cell(r, c).value = comps[0].formula
    r, c = parse_cell_ref(comps[1].cell)
    wb[comps[1].tab].cell(r, c).value = "=1+1"
    wb.save(trainer_path)
    wb.close()

    main(["check", "--workbook", str(trainer_path)])
    out = capsys.readouterr().out

    for comp in comps:
        assert comp.formula not in out
        if comp.short_hint:
            assert comp.short_hint not in out
        for h in comp.hints:
            assert h not in out
        if comp.expected_value is not None:
            assert str(comp.expected_value) not in out


def test_cli_list_resolved(tmp_path):
    trainer_path, _ = _build_pair(tmp_path)
    from core.__main__ import main

    assert main(["list", "--workbook", str(trainer_path)]) == 0


def test_cli_build_reports_both_paths(tmp_path, capsys):
    from core.__main__ import main

    out = tmp_path / "DEMO_HK_Trainer.xlsx"
    rc = main(["build", str(DEMO_JSON), "-o", str(out)])
    captured = capsys.readouterr()
    assert rc == 0
    assert "Trainer workbook:" in captured.out
    assert "Answer Key workbook:" in captured.out
    assert (tmp_path / "DEMO_HK_Trainer.xlsx").exists()
    assert (tmp_path / "DEMO_HK_Answer_Key.xlsx").exists()


def test_hint_reveal_modules_removed():
    assert importlib.util.find_spec("core.trainer.hints") is None
    assert not (ROOT / "core" / "templates" / "TrainerMacros.bas").exists()


_SSML = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_RELS = "http://schemas.openxmlformats.org/package/2006/relationships"
_OD_RELS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _xlsx_sheet_path(zf, sheet_name: str) -> str:
    import xml.etree.ElementTree as ET

    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    sheets = wb.find(f"{{{_SSML}}}sheets")
    if sheets is None:
        raise FileNotFoundError("workbook.xml missing sheets")
    rid = None
    for sheet in sheets:
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib.get(f"{{{_OD_RELS}}}id")
            break
    if not rid:
        raise FileNotFoundError(f"Sheet not found: {sheet_name}")
    target = None
    for rel in rels:
        if rel.attrib.get("Id") == rid:
            target = rel.attrib.get("Target", "")
            break
    if not target:
        raise FileNotFoundError(f"Relationship missing for sheet {sheet_name}")
    target = target.lstrip("/")
    if not target.startswith("xl/"):
        target = f"xl/{target}"
    return target


def _inject_formula_and_cached_value(
    workbook_path: Path,
    sheet: str,
    cell: str,
    *,
    formula: str,
    cached_value: float,
) -> None:
    """Patch one cell's OOXML <f> and cached <v> without openpyxl rewrite."""
    import os
    import tempfile
    import zipfile
    import xml.etree.ElementTree as ET

    formula_body = formula[1:] if formula.startswith("=") else formula
    workbook_path = Path(workbook_path)

    with zipfile.ZipFile(workbook_path, "r") as zf_in:
        sheet_path = _xlsx_sheet_path(zf_in, sheet)
        sheet_xml = zf_in.read(sheet_path)
        other_members = {
            info.filename: zf_in.read(info.filename)
            for info in zf_in.infolist()
            if info.filename != sheet_path
        }

    root = ET.fromstring(sheet_xml)
    cell_el = None
    for c_el in root.iter(f"{{{_SSML}}}c"):
        if c_el.attrib.get("r") == cell:
            cell_el = c_el
            break
    if cell_el is None:
        raise FileNotFoundError(f"Cell {cell} not found on sheet {sheet}")

    # Replace formula / cached value children; keep style and address.
    for child in list(cell_el):
        if child.tag in {f"{{{_SSML}}}f", f"{{{_SSML}}}v", f"{{{_SSML}}}is"}:
            cell_el.remove(child)
    f_el = ET.SubElement(cell_el, f"{{{_SSML}}}f")
    f_el.text = formula_body
    v_el = ET.SubElement(cell_el, f"{{{_SSML}}}v")
    v_el.text = f"{float(cached_value)}"

    ET.register_namespace("", _SSML)
    new_sheet_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    fd, tmp_name = tempfile.mkstemp(
        suffix=".xlsx", prefix=workbook_path.stem + "_inj_", dir=str(workbook_path.parent)
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        with zipfile.ZipFile(tmp_path, "w") as zf_out:
            for name, data in other_members.items():
                zf_out.writestr(name, data)
            zf_out.writestr(sheet_path, new_sheet_xml)
        tmp_path.replace(workbook_path)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def test_equivalent_cached_formula_stays_correct_across_repeated_checks(tmp_path):
    trainer_path, answer_key_path = _build_pair(tmp_path)
    smap = load_semantic_map(answer_key_path)
    comp = next(
        c for c in smap.all_ordered() if isinstance(c.expected_value, (int, float))
    )

    _inject_formula_and_cached_value(
        trainer_path,
        comp.tab,
        comp.cell,
        formula=f"={float(comp.expected_value)}",
        cached_value=float(comp.expected_value),
    )

    wb_formula = load_workbook(trainer_path, data_only=False)
    row, col = parse_cell_ref(comp.cell)
    entered_formula = wb_formula[comp.tab].cell(row, col).value
    wb_formula.close()
    assert entered_formula != comp.formula

    first = check_workbook(trainer_path)
    assert first.correct >= 1

    # Check itself must not destroy the cached Excel result.
    wb_cached = load_workbook(trainer_path, data_only=True)
    cached_after_first_check = wb_cached[comp.tab].cell(row, col).value
    wb_cached.close()
    assert cached_after_first_check == pytest.approx(float(comp.expected_value))

    second = check_workbook(trainer_path)
    assert second.correct >= 1

    wb = load_workbook(trainer_path, data_only=False)
    assert _fill_rgb(wb[comp.tab].cell(row, col)) == "C8E6C9"
    assert wb[comp.tab].cell(row, col).value == entered_formula
    wb.close()
